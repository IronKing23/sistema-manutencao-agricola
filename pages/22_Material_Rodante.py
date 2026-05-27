import streamlit as st
import sqlite3
import pandas as pd
from datetime import datetime
import json
import sys
import os
import io
import base64
import time
from fpdf import FPDF
import tempfile

# Tenta importar openpyxl para gerar o Espelho em Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# --- BLINDAGEM DE IMPORTAÇÃO E UI ---
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

try:
    from database import get_db_connection
    from utils_ui import load_custom_css, ui_header, ui_empty_state
except ImportError:
    def get_db_connection():
        return sqlite3.connect('manutencao.db', check_same_thread=False)


    def load_custom_css():
        pass


    def ui_header(title, subtitle, icon):
        st.title(f"{icon} {title}"); st.caption(subtitle)


    def ui_empty_state(msg, icon):
        st.info(f"{icon} {msg}")

load_custom_css()
ui_header("Inspeção de Material Rodante",
          "Coleta de dados digitalizada, fotos, edição de histórico e relatórios PDF/Excel.", "🚜")


# ==============================================================================
# BASE DE DADOS
# ==============================================================================
def inicializar_tabela_rodante():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS inspecao_rodante (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data_inspecao TEXT,
            tecnico TEXT,
            local TEXT,
            frota TEXT,
            modelo TEXT,
            horimetro REAL,
            agreg_le TEXT,
            agreg_ld TEXT,
            dados_json TEXT
        )
    """)
    conn.commit()
    conn.close()


inicializar_tabela_rodante()


def carregar_frotas():
    try:
        conn = get_db_connection()
        frotas = pd.read_sql_query("SELECT frota, modelo FROM equipamentos ORDER BY frota", conn)
        conn.close()
        return frotas
    except:
        return pd.DataFrame(columns=['frota', 'modelo'])


def carregar_historico():
    conn = get_db_connection()
    df = pd.read_sql_query("SELECT id, data_inspecao, frota, tecnico, horimetro FROM inspecao_rodante ORDER BY id DESC",
                           conn)
    conn.close()
    return df


def buscar_inspecao_completa(id_inspecao):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM inspecao_rodante WHERE id = ?", (id_inspecao,))
    row = cursor.fetchone()
    columns = [col[0] for col in cursor.description] if cursor.description else []
    conn.close()
    return dict(zip(columns, row)) if row else None


def deletar_inspecao(id_inspecao):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM inspecao_rodante WHERE id = ?", (id_inspecao,))
    conn.commit()
    conn.close()


# ==============================================================================
# HELPERS DE FORMULÁRIO
# ==============================================================================
def get_index(opcoes, valor):
    """Helper para encontrar o índice seguro de um selectbox baseado em valor anterior"""
    if not opcoes or valor not in opcoes:
        return 0
    return opcoes.index(valor)


def render_linha_medida(nome_item, id_item, prefixo, opcoes_marca, opcoes_medida, opcoes_condicao, defaults, id_form):
    # Tratamento de retrocompatibilidade para vistorias antigas
    if isinstance(defaults, list):
        defaults = {'condicao': defaults}
    elif not isinstance(defaults, dict):
        defaults = {}

    st.markdown(f"**{nome_item}**")
    row_data = {}
    c_marca, c_cond, c_medidas = st.columns([1.5, 2, 2])

    with c_marca:
        if opcoes_marca:
            idx = get_index(opcoes_marca, defaults.get('marca'))
            row_data['marca'] = st.selectbox("Marca/Tipo", opcoes_marca, index=idx,
                                             key=f"{prefixo}{id_item}_{id_form}_marca")
        if opcoes_medida:
            idx = get_index(opcoes_medida, defaults.get('medida_padrao'))
            row_data['medida_padrao'] = st.selectbox("Padrão", opcoes_medida, index=idx,
                                                     key=f"{prefixo}{id_item}_{id_form}_medida")

    with c_cond:
        if opcoes_condicao:
            def_cond = defaults.get('condicao', [])
            def_cond = [c for c in def_cond if c in opcoes_condicao]  # Segurança
            row_data['condicao'] = st.multiselect("Anomalias/Condição", opcoes_condicao, default=def_cond,
                                                  key=f"{prefixo}{id_item}_{id_form}_cond")

    with c_medidas:
        mc1, mc2, mc3 = st.columns(3)
        row_data['std'] = mc1.number_input("STD", min_value=0.0, step=0.1, value=float(defaults.get('std', 0.0)),
                                           key=f"{prefixo}{id_item}_{id_form}_std")
        row_data['atual'] = mc2.number_input("Atual", min_value=0.0, step=0.1, value=float(defaults.get('atual', 0.0)),
                                             key=f"{prefixo}{id_item}_{id_form}_atual")
        row_data['restante'] = mc3.number_input("% Rest.", min_value=0.0, max_value=100.0, step=1.0,
                                                value=float(defaults.get('restante', 0.0)),
                                                key=f"{prefixo}{id_item}_{id_form}_rest")

    return row_data


def renderizar_lado_rodante(prefixo, defaults, id_form):
    dados = {}
    st.markdown(f"#### ⚙️ Avaliação Geral - {prefixo.replace('_', '')}")
    col1, col2 = st.columns(2)

    with col1:
        opts_truck = ["OK", "Alinhado", "Empenado", "Trincado", "Desalinhado"]
        dados['truck_status'] = st.selectbox("Status do Truck", opts_truck,
                                             index=get_index(opts_truck, defaults.get('truck_status')),
                                             key=f"{prefixo}{id_form}_truck_st")

        opts_comp = ["Bola", "Eixo", "Rótula"]
        dados['truck_comp'] = st.multiselect("Componentes do Truck", opts_comp,
                                             default=[c for c in defaults.get('truck_comp', []) if c in opts_comp],
                                             key=f"{prefixo}{id_form}_truck_cp")

        opts_prot = ["OK", "Não Possui", "Desgastada"]
        dados['protecao_rolete'] = st.selectbox("Proteção do Rolete", opts_prot,
                                                index=get_index(opts_prot, defaults.get('protecao_rolete')),
                                                key=f"{prefixo}{id_form}_prot_rol")

    with col2:
        opts_guia = ["OK", "Faltando", "Suporte Danificado"]
        dados['guia_esteira_status'] = st.selectbox("Guia da Esteira", opts_guia,
                                                    index=get_index(opts_guia, defaults.get('guia_esteira_status')),
                                                    key=f"{prefixo}{id_form}_guia_st")
        dados['guia_esteira_qtd'] = st.number_input("Quant. Guias Danificadas", min_value=0, step=1,
                                                    value=int(defaults.get('guia_esteira_qtd', 0)),
                                                    key=f"{prefixo}{id_form}_guia_qtd")

        opts_mao = ["OK", "Soldada", "Sem Aperto", "Paraf. Quebrado"]
        dados['mao_amigo'] = st.multiselect("Mão de Amigo", opts_mao,
                                            default=[c for c in defaults.get('mao_amigo', []) if c in opts_mao],
                                            key=f"{prefixo}{id_form}_mao_amg")

    st.markdown("---")
    st.markdown(f"#### 📏 Medições e Desgastes - {prefixo.replace('_', '')}")

    dados['elo'] = render_linha_medida("Elo", "elo", prefixo, ["ITM", "BRC", "VTK", "ITR"], ["96", "98", "106"],
                                       ["Trincado", "Lascado", "Desgaste Lat."], defaults.get('elo', {}), id_form)
    dados['bucha'] = render_linha_medida("Bucha", "bucha", prefixo, ["Seca", "Lubrificada", "Graxa"],
                                         ["53.8", "57", "60"], ["Girada", "Novo", "Trincado/Furado"],
                                         defaults.get('bucha', {}), id_form)
    dados['passo'] = render_linha_medida("Passo", "passo", prefixo, None, None,
                                         ["T. Correta", "T. Excessiva", "T. Frouxa"], defaults.get('passo', {}),
                                         id_form)
    dados['sapata'] = render_linha_medida("Sapata", "sapata", prefixo, ["Paralela", "Trapezoidal"], None,
                                          ["Novo", "Recuperado", "Desgastado"], defaults.get('sapata', {}), id_form)
    dados['roda_guia'] = render_linha_medida("Roda Guia", "roda_guia", prefixo, ["Novo", "Recondicionado"],
                                             ["17", "21", "23"], ["Folga", "Desgaste Mancal"],
                                             defaults.get('roda_guia', {}), id_form)
    dados['roda_motriz'] = render_linha_medida("Roda Motriz", "roda_motriz", prefixo, ["ITM", "BRC", "VTK", "ITR"],
                                               None,
                                               ["Desg. Ponta", "Desg. Lateral", "Paraf. Quebrado", "Paraf. Faltando"],
                                               defaults.get('roda_motriz', {}), id_form)
    dados['rolete_sup'] = render_linha_medida("Rolete Superior", "rolete_sup", prefixo, None, None,
                                              ["Novo", "Vazamento", "Travado", "Recond.", "Desborrachado"],
                                              defaults.get('rolete_sup', {}), id_form)

    st.markdown("---")
    st.markdown("#### 🛞 Roletes Inferiores")
    for i in range(1, 9):
        dados[f'rolete_inf_{i}'] = render_linha_medida(
            f"Nº {i}", f"rolete_inf_{i}", prefixo,
            None, None, ["Normal", "Desabou", "Vazamento"],
            defaults.get(f'rolete_inf_{i}', {}), id_form
        )

    return dados


# ==============================================================================
# MOTOR DO FORMULÁRIO (USADO PARA NOVO E EDITAR)
# ==============================================================================
def render_form_vistoria(registro_bd=None):
    modo_edicao = registro_bd is not None
    id_form = registro_bd['id'] if modo_edicao else "novo"

    # Extrai JSON
    json_data = {}
    if modo_edicao and registro_bd.get('dados_json'):
        json_data = json.loads(registro_bd['dados_json'])

    def_le = json_data.get('LE', {})
    def_ld = json_data.get('LD', {})
    def_obs = json_data.get('observacoes', '')
    def_fotos_antigas = json_data.get('fotos', [])

    with st.form(f"form_rodante_{id_form}"):
        st.subheader("1. Identificação do Equipamento")
        c1, c2, c3 = st.columns(3)

        with c1:
            nome_tecnico = st.text_input("Técnico Responsável*", value=registro_bd['tecnico'] if modo_edicao else "",
                                         placeholder="Seu nome", key=f"tec_{id_form}")

            dt_val = datetime.today()
            if modo_edicao and registro_bd['data_inspecao']:
                dt_val = datetime.strptime(registro_bd['data_inspecao'], '%Y-%m-%d').date()
            data_coleta = st.date_input("Data da Avaliação*", value=dt_val, key=f"dt_{id_form}")

            local = st.text_input("Local / Usina*", value=registro_bd['local'] if modo_edicao else "Usina Cedro",
                                  key=f"loc_{id_form}")

        with c2:
            df_frotas = carregar_frotas()
            frotas_lista = df_frotas['frota'].tolist() if not df_frotas.empty else []

            idx_frota = None
            if modo_edicao and registro_bd['frota'] in frotas_lista:
                idx_frota = frotas_lista.index(registro_bd['frota'])

            frota_selecionada = st.selectbox("Frota*", options=frotas_lista, index=idx_frota,
                                             placeholder="Selecione a Frota", key=f"frota_{id_form}")

            modelo_maquina = registro_bd['modelo'] if modo_edicao else ""
            if frota_selecionada and not df_frotas.empty and not modo_edicao:
                modelo_maquina = df_frotas[df_frotas['frota'] == frota_selecionada]['modelo'].values[0]

            modelo = st.text_input("Modelo", value=modelo_maquina, key=f"mod_{id_form}")
            horimetro = st.number_input("Horímetro", min_value=0.0, step=1.0,
                                        value=float(registro_bd['horimetro']) if modo_edicao else 0.0,
                                        key=f"hor_{id_form}")

        with c3:
            agreg_le = st.text_input("Nº Agregado L.E.", value=registro_bd['agreg_le'] if modo_edicao else "",
                                     key=f"agle_{id_form}")
            agreg_ld = st.text_input("Nº Agregado L.D.", value=registro_bd['agreg_ld'] if modo_edicao else "",
                                     key=f"agld_{id_form}")

        st.markdown("---")
        st.subheader("2. Avaliação de Lados")

        t_le, t_ld = st.tabs(["⬅️ Lado Esquerdo (L.E.)", "➡️ Lado Direito (L.D.)"])
        with t_le:
            dados_le = renderizar_lado_rodante("LE_", def_le, id_form)
        with t_ld:
            dados_ld = renderizar_lado_rodante("LD_", def_ld, id_form)

        st.markdown("---")
        st.subheader("3. Observações e Evidências Fotográficas")

        observacoes_vistoria = st.text_area("Observações Gerais sobre a Vistoria", value=def_obs,
                                            placeholder="Anote aqui qualquer detalhe...", key=f"obs_{id_form}")

        if modo_edicao and def_fotos_antigas:
            st.info(
                f"📸 Esta inspeção já possui {len(def_fotos_antigas)} foto(s) anexada(s). Novas fotos serão adicionadas a elas.")
            limpar_fotos = st.checkbox("🗑️ Apagar fotos antigas do banco de dados ao salvar", value=False,
                                       key=f"limpar_{id_form}")
        else:
            limpar_fotos = False

        st.info("💡 **Dica de Câmera:** Se estiver no telemóvel, clicar em 'Browse files' permitirá usar a câmera.")
        fotos_upload = st.file_uploader("📷 Anexar Novas Fotos", accept_multiple_files=True, type=['png', 'jpg', 'jpeg'],
                                        key=f"up_fotos_{id_form}")

        st.markdown("---")
        texto_botao = "💾 Atualizar Inspeção Existente" if modo_edicao else "💾 Salvar Nova Inspeção"
        btn_salvar = st.form_submit_button(texto_botao, type="primary", use_container_width=True)

        if btn_salvar:
            if not nome_tecnico or not frota_selecionada:
                st.error("Preencha o Nome do Técnico e a Frota!")
            else:
                fotos_finais = [] if limpar_fotos else def_fotos_antigas.copy()
                if fotos_upload:
                    for foto in fotos_upload:
                        b64 = base64.b64encode(foto.read()).decode('utf-8')
                        fotos_finais.append(b64)

                payload_json = {
                    "LE": dados_le,
                    "LD": dados_ld,
                    "observacoes": observacoes_vistoria,
                    "fotos": fotos_finais
                }

                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()

                    if modo_edicao:
                        cursor.execute("""
                            UPDATE inspecao_rodante SET 
                                data_inspecao=?, tecnico=?, local=?, frota=?, modelo=?, 
                                horimetro=?, agreg_le=?, agreg_ld=?, dados_json=?
                            WHERE id=?
                        """, (
                            data_coleta.strftime('%Y-%m-%d'), nome_tecnico, local,
                            frota_selecionada, modelo, horimetro, agreg_le, agreg_ld,
                            json.dumps(payload_json), id_form
                        ))
                        msg = "✅ Atualização guardada com sucesso!"
                    else:
                        cursor.execute("""
                            INSERT INTO inspecao_rodante 
                            (data_inspecao, tecnico, local, frota, modelo, horimetro, agreg_le, agreg_ld, dados_json)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            data_coleta.strftime('%Y-%m-%d'), nome_tecnico, local,
                            frota_selecionada, modelo, horimetro, agreg_le, agreg_ld,
                            json.dumps(payload_json)
                        ))
                        msg = "✅ Nova inspeção criada com sucesso!"

                    conn.commit()
                    conn.close()
                    st.success(msg)
                    st.balloons()
                    time.sleep(1.5)
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao salvar na base de dados: {e}")


# ==============================================================================
# GERAÇÃO DE EXCEL (ESPELHO)
# ==============================================================================
def gerar_excel_rodante(registro):
    if not OPENPYXL_AVAILABLE: return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Avaliação Material Rodante"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    sub_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    bold = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))
    align_c = Alignment(horizontal="center", vertical="center")
    align_l = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells('A1:F1')
    ws['A1'] = "AVALIAÇÃO DE MATERIAL RODANTE - USINA CEDRO"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = align_c

    dt_obj = datetime.strptime(registro['data_inspecao'], '%Y-%m-%d')
    dt_br = dt_obj.strftime('%d/%m/%Y')

    ws.append(["Técnico:", registro['tecnico'], "", "Data:", dt_br])
    ws.append(["Frota/Modelo:", f"{registro['frota']} / {registro['modelo']}", "", "Horímetro:", registro['horimetro']])
    ws.append(["Agregado L.E.:", registro['agreg_le'], "", "Agregado L.D.:", registro['agreg_ld']])
    ws.append([])

    for row in range(2, 5):
        ws.cell(row=row, column=1).font = bold
        ws.cell(row=row, column=4).font = bold

    dados = json.loads(registro.get('dados_json', '{}'))

    def preencher_lado(titulo, lado_data):
        ws.append([titulo])
        tit_cell = ws.cell(row=ws.max_row, column=1)
        tit_cell.font = bold
        tit_cell.fill = sub_fill
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)

        ws.append(["TRUCK", "GUIA ESTEIRA", "PROT. ROLETE", "MÃO DE AMIGO", "", ""])
        for i in range(1, 5):
            c = ws.cell(row=ws.max_row, column=i)
            c.font = bold
            c.alignment = align_c

        t_st = str(lado_data.get('truck_status', ''))
        t_cp = ", ".join(lado_data.get('truck_comp', []))
        g_st = str(lado_data.get('guia_esteira_status', ''))
        g_qt = str(lado_data.get('guia_esteira_qtd', '0'))
        p_st = str(lado_data.get('protecao_rolete', ''))
        m_am = ", ".join(lado_data.get('mao_amigo', []))

        ws.append([f"{t_st} | {t_cp}", f"{g_st} (Qtd: {g_qt})", p_st, m_am, "", ""])
        for i in range(1, 5): ws.cell(row=ws.max_row, column=i).alignment = align_c
        ws.append([])

        head_row = ["COMPONENTE", "ESPECIFICAÇÕES", "ANOMALIAS / CONDIÇÕES", "STD", "ATUAL", "% REST."]
        ws.append(head_row)
        for i in range(1, 7):
            c = ws.cell(row=ws.max_row, column=i)
            c.font = header_font
            c.fill = header_fill
            c.alignment = align_c
            c.border = border

        itens = [
            ('ELO', 'elo'), ('BUCHA', 'bucha'), ('PASSO', 'passo'),
            ('SAPATA', 'sapata'), ('RODA GUIA', 'roda_guia'),
            ('RODA MOTRIZ', 'roda_motriz'), ('ROLETE SUP.', 'rolete_sup')
        ]
        # Inserir os roletes inferiores no fluxo principal
        for i in range(1, 9):
            itens.append((f'ROLETE INF. {i}', f'rolete_inf_{i}'))

        for nome, chave in itens:
            item_data = lado_data.get(chave, {})
            # Tratamento de retrocompatibilidade para vistorias antigas
            if isinstance(item_data, list):
                item_data = {'condicao': item_data}
            elif not isinstance(item_data, dict):
                item_data = {}

            esp = []
            if item_data.get('marca'): esp.append(item_data['marca'])
            if item_data.get('medida_padrao'): esp.append(item_data['medida_padrao'])
            str_esp = " / ".join(esp)
            str_cond = ", ".join(item_data.get('condicao', []))
            std = str(item_data.get('std', 0))
            atual = str(item_data.get('atual', 0))
            rest = str(item_data.get('restante', 0)) + "%"

            ws.append([nome, str_esp, str_cond, std, atual, rest])
            for i in range(1, 7):
                c = ws.cell(row=ws.max_row, column=i)
                c.border = border
                if i > 3: c.alignment = align_c
        ws.append([])

    preencher_lado("LADO ESQUERDO (L.E.)", dados.get('LE', {}))
    preencher_lado("LADO DIREITO (L.D.)", dados.get('LD', {}))

    ws.append(["OBSERVAÇÕES GERAIS:"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([dados.get('observacoes', 'Nenhuma observação registada.')])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
    ws.cell(row=ws.max_row, column=1).alignment = align_l

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ==============================================================================
# GERAÇÃO DE PDF (COM ESCUDO ANTI-ERROS E IMAGENS)
# ==============================================================================
class PDFMaterialRodante(FPDF):
    def _sanitize(self, text):
        if text is None: return ""
        return str(text).encode('latin-1', 'ignore').decode('latin-1')

    def cell(self, *args, **kwargs):
        new_args = list(args)
        if len(new_args) >= 3: new_args[2] = self._sanitize(new_args[2])
        if 'txt' in kwargs: kwargs['txt'] = self._sanitize(kwargs['txt'])
        super().cell(*new_args, **kwargs)

    def multi_cell(self, *args, **kwargs):
        new_args = list(args)
        if len(new_args) >= 3: new_args[2] = self._sanitize(new_args[2])
        if 'txt' in kwargs: kwargs['txt'] = self._sanitize(kwargs['txt'])
        super().multi_cell(*new_args, **kwargs)

    def header(self):
        caminho_logo = os.path.join(os.path.dirname(os.path.dirname(__file__)), "logo_cedro.png")
        if os.path.exists(caminho_logo):
            self.image(caminho_logo, 10, 8, 15)

        self.set_font('Arial', 'B', 14)
        self.cell(0, 8, 'AVALIACAO DE MATERIAL RODANTE', 0, 1, 'C')
        self.set_font('Arial', '', 9)
        self.cell(0, 4, 'Usina Cedro - Pedra Agroindustrial S/A', 0, 1, 'C')
        self.ln(5)


def gerar_pdf_rodante(registro):
    pdf = PDFMaterialRodante()
    pdf.add_page()

    pdf.set_font('Arial', 'B', 9)
    pdf.set_fill_color(240, 240, 240)

    pdf.cell(30, 6, 'Tecnico:', 1, 0, 'L', fill=True)
    pdf.set_font('Arial', '', 9)
    pdf.cell(65, 6, str(registro['tecnico']), 1, 0, 'L')

    pdf.set_font('Arial', 'B', 9)
    pdf.cell(30, 6, 'Data:', 1, 0, 'L', fill=True)
    pdf.set_font('Arial', '', 9)
    dt_obj = datetime.strptime(registro['data_inspecao'], '%Y-%m-%d')
    pdf.cell(65, 6, dt_obj.strftime('%d/%m/%Y'), 1, 1, 'L')

    pdf.set_font('Arial', 'B', 9)
    pdf.cell(30, 6, 'Frota / Modelo:', 1, 0, 'L', fill=True)
    pdf.set_font('Arial', '', 9)
    pdf.cell(65, 6, f"{registro['frota']} / {registro['modelo']}", 1, 0, 'L')

    pdf.set_font('Arial', 'B', 9)
    pdf.cell(30, 6, 'Horimetro:', 1, 0, 'L', fill=True)
    pdf.set_font('Arial', '', 9)
    pdf.cell(65, 6, str(registro['horimetro']), 1, 1, 'L')

    pdf.set_font('Arial', 'B', 9)
    pdf.cell(30, 6, 'Agregado L.E:', 1, 0, 'L', fill=True)
    pdf.set_font('Arial', '', 9)
    pdf.cell(65, 6, str(registro['agreg_le']), 1, 0, 'L')

    pdf.set_font('Arial', 'B', 9)
    pdf.cell(30, 6, 'Agregado L.D:', 1, 0, 'L', fill=True)
    pdf.set_font('Arial', '', 9)
    pdf.cell(65, 6, str(registro['agreg_ld']), 1, 1, 'L')

    pdf.ln(5)

    dados = json.loads(registro.get('dados_json', '{}'))

    def imprimir_lado(titulo, lado_data):
        pdf.set_font('Arial', 'B', 11)
        pdf.set_fill_color(200, 200, 200)
        pdf.cell(0, 7, titulo, 1, 1, 'C', fill=True)

        pdf.set_font('Arial', 'B', 8)
        pdf.cell(47.5, 5, 'TRUCK', 1, 0, 'C')
        pdf.cell(47.5, 5, 'GUIA ESTEIRA', 1, 0, 'C')
        pdf.cell(47.5, 5, 'PROT. ROLETE', 1, 0, 'C')
        pdf.cell(47.5, 5, 'MAO DE AMIGO', 1, 1, 'C')

        pdf.set_font('Arial', '', 7)
        t_st = str(lado_data.get('truck_status', ''))
        t_cp = ", ".join(lado_data.get('truck_comp', []))
        g_st = str(lado_data.get('guia_esteira_status', ''))
        g_qt = str(lado_data.get('guia_esteira_qtd', '0'))
        p_st = str(lado_data.get('protecao_rolete', ''))
        m_am = ", ".join(lado_data.get('mao_amigo', []))

        pdf.cell(47.5, 5, f"{t_st} | {t_cp}", 1, 0, 'C')
        pdf.cell(47.5, 5, f"{g_st} (Qtd: {g_qt})", 1, 0, 'C')
        pdf.cell(47.5, 5, f"{p_st}", 1, 0, 'C')
        pdf.cell(47.5, 5, f"{m_am}", 1, 1, 'C')

        pdf.ln(2)

        pdf.set_font('Arial', 'B', 8)
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(30, 5, 'COMPONENTE', 1, 0, 'C', fill=True)
        pdf.cell(50, 5, 'ESPECIFICACOES', 1, 0, 'C', fill=True)
        pdf.cell(70, 5, 'ANOMALIAS / CONDICOES', 1, 0, 'C', fill=True)
        pdf.cell(13, 5, 'STD', 1, 0, 'C', fill=True)
        pdf.cell(13, 5, 'ATUAL', 1, 0, 'C', fill=True)
        pdf.cell(14, 5, '% REST', 1, 1, 'C', fill=True)

        pdf.set_font('Arial', '', 7)
        itens = [
            ('ELO', 'elo'), ('BUCHA', 'bucha'), ('PASSO', 'passo'),
            ('SAPATA', 'sapata'), ('RODA GUIA', 'roda_guia'),
            ('RODA MOTRIZ', 'roda_motriz'), ('ROLETE SUP.', 'rolete_sup')
        ]
        # Adiciona roletes inferiores no PDF
        for i in range(1, 9):
            itens.append((f'ROLETE INF. {i}', f'rolete_inf_{i}'))

        for nome_display, chave in itens:
            item_data = lado_data.get(chave, {})
            # Tratamento de retrocompatibilidade para vistorias antigas
            if isinstance(item_data, list):
                item_data = {'condicao': item_data}
            elif not isinstance(item_data, dict):
                item_data = {}

            esp = []
            if item_data.get('marca'): esp.append(item_data['marca'])
            if item_data.get('medida_padrao'): esp.append(item_data['medida_padrao'])
            str_esp = " / ".join(esp)

            str_cond = ", ".join(item_data.get('condicao', []))

            pdf.cell(30, 5, nome_display, 1, 0, 'L')
            pdf.cell(50, 5, str_esp[:35], 1, 0, 'L')
            pdf.cell(70, 5, str_cond[:50], 1, 0, 'L')
            pdf.cell(13, 5, str(item_data.get('std', 0)), 1, 0, 'C')
            pdf.cell(13, 5, str(item_data.get('atual', 0)), 1, 0, 'C')
            pdf.cell(14, 5, str(item_data.get('restante', 0)) + "%", 1, 1, 'C')

        pdf.ln(5)

    imprimir_lado("LADO ESQUERDO (L.E.)", dados.get('LE', {}))
    imprimir_lado("LADO DIREITO (L.D.)", dados.get('LD', {}))

    obs = dados.get('observacoes', '')
    if obs:
        pdf.ln(2)
        pdf.set_font('Arial', 'B', 9)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(0, 6, 'OBSERVACOES GERAIS:', 1, 1, 'L', fill=True)
        pdf.set_font('Arial', '', 8)
        pdf.multi_cell(0, 5, obs, 1, 'L')

    fotos = dados.get('fotos', [])
    if fotos:
        pdf.add_page()
        pdf.set_font('Arial', 'B', 12)
        pdf.set_fill_color(200, 200, 200)
        pdf.cell(0, 10, 'ANEXO FOTOGRAFICO', 1, 1, 'C', fill=True)
        pdf.ln(5)

        x_positions = [10, 105]
        col_idx = 0
        y_curr = pdf.get_y()
        img_height = 70

        for b64_str in fotos:
            foto_bytes = base64.b64decode(b64_str)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp_foto:
                tmp_foto.write(foto_bytes)
                tmp_path = tmp_foto.name

            if col_idx == 2:
                col_idx = 0
                y_curr += img_height + 5
                if y_curr + img_height > 280:
                    pdf.add_page()
                    y_curr = pdf.get_y()
            try:
                pdf.image(tmp_path, x=x_positions[col_idx], y=y_curr, w=90)
            except Exception as e:
                pass
            finally:
                os.remove(tmp_path)

            col_idx += 1

    tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    pdf_path = tmp_pdf.name
    tmp_pdf.close()

    pdf.output(pdf_path)

    with open(pdf_path, "rb") as f:
        bytes_pdf = f.read()

    os.remove(pdf_path)
    return bytes_pdf


# ==============================================================================
# INTERFACE PRINCIPAL - TABS
# ==============================================================================
tab_nova, tab_historico = st.tabs(["📝 Nova Inspeção", "📋 Histórico / Consultar / Editar"])

with tab_nova:
    st.info("Preencha os campos abaixo para registrar uma nova inspeção.")
    render_form_vistoria()  # Chama o motor sem registro (modo NOVO)

with tab_historico:
    st.subheader("Consultar, Editar e Exportar")
    st.caption("Filtre as inspeções, carregue os dados para alterá-los ou gere o PDF/Excel.")

    df_hist = carregar_historico()

    if df_hist.empty:
        ui_empty_state("Nenhuma inspeção de material rodante registrada ainda.", "📄")
    else:
        # Filtros
        c_f1, c_f2 = st.columns(2)
        with c_f1:
            frotas_cad = ["Todas as Frotas"] + sorted(df_hist['frota'].unique().tolist())
            filtro_frota = st.selectbox("Filtrar por Frota", options=frotas_cad)
        with c_f2:
            filtro_data = st.date_input("Filtrar por Período", value=[])

        df_filtrado = df_hist.copy()
        if filtro_frota != "Todas as Frotas":
            df_filtrado = df_filtrado[df_filtrado['frota'] == filtro_frota]
        if len(filtro_data) == 2:
            dt_inicio = filtro_data[0].strftime('%Y-%m-%d')
            dt_fim = filtro_data[1].strftime('%Y-%m-%d')
            df_filtrado = df_filtrado[
                (df_filtrado['data_inspecao'] >= dt_inicio) & (df_filtrado['data_inspecao'] <= dt_fim)]

        df_filtrado['display'] = "ID " + df_filtrado['id'].astype(str) + " - " + df_filtrado['frota'] + " (" + \
                                 df_filtrado['data_inspecao'] + ")"

        st.markdown("---")
        inspecao_sel = st.selectbox("Selecione a Avaliação para Consultar/Editar:", df_filtrado['display'], index=None)

        if inspecao_sel:
            id_banco = int(inspecao_sel.split(" - ")[0].replace("ID ", ""))
            registro_completo = buscar_inspecao_completa(id_banco)

            if registro_completo:
                # --- BOTÕES DE AÇÃO ---
                col_btn1, col_btn2, col_btn3 = st.columns([2, 2, 1])

                with col_btn1:
                    pdf_bytes = gerar_pdf_rodante(registro_completo)
                    st.download_button(
                        label="🖨️ Exportar PDF",
                        data=pdf_bytes,
                        file_name=f"Inspecao_Rodante_{registro_completo['frota']}_{registro_completo['data_inspecao']}.pdf",
                        mime="application/pdf",
                        type="primary",
                        use_container_width=True
                    )

                with col_btn2:
                    if OPENPYXL_AVAILABLE:
                        excel_bytes = gerar_excel_rodante(registro_completo)
                        if excel_bytes:
                            st.download_button(
                                label="📊 Exportar Excel",
                                data=excel_bytes,
                                file_name=f"Inspecao_Rodante_{registro_completo['frota']}_{registro_completo['data_inspecao']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                with col_btn3:
                    if st.button("🗑️ Excluir", use_container_width=True):
                        deletar_inspecao(id_banco)
                        st.success("Excluído com sucesso!")
                        st.rerun()

                st.markdown("---")
                st.markdown("### ✏️ Formulário de Edição")
                st.info(
                    f"Visualizando os dados do **Técnico:** {registro_completo['tecnico']} | **Horímetro:** {registro_completo['horimetro']}")

                # Renderiza o mesmo formulário, mas preenchido com os dados do banco!
                render_form_vistoria(registro_completo)