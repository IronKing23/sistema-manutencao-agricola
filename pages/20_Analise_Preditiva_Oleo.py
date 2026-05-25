import streamlit as st
import pandas as pd
import numpy as np
import sys
import os
import unicodedata
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from fpdf import FPDF
import tempfile
import io
import time
import glob
import logging
from contextlib import contextmanager

# Logger para substituir excepts silenciosos
logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger("oleo_preditivo")

# ==============================================================================
# LIMITES DE REFERÊNCIA POR ELEMENTO (constante global)
# Usados pelo Health Score, pelo PDF colorido e pela UI Streamlit.
# Valores típicos para óleo de motor — ajustar por compartimento se necessário.
# Formato: elemento → (verde_ate, amarelo_ate) → acima = vermelho
# ==============================================================================
LIMITES_REFERENCIA = {
    'Ferro':      (50,  100),
    'Cromo':      (5,   15),
    'Níquel':     (3,   10),
    'Chumbo':     (10,  30),
    'Estanho':    (5,   15),
    'Cobre':      (15,  40),
    'Alumínio':   (10,  25),
    'Manganês':   (3,   8),
    'Silício':    (15,  30),
    'Sódio':      (20,  50),
    'Potássio':   (15,  40),
    'Boro':       (50,  200),
    'INDICE_PQ':  (25,  50),
    'Diluição Diesel': (2.0, 4.0),
    'TBN':           (10, 5),     # invertido: ABAIXO de 5 é crítico
    'KF_Agua':       (200, 500),
    'Oxidação_FTIR':   (15, 25),
    'Nitração_FTIR':   (15, 25),
    'Sulfatação_FTIR': (20, 35),
    'Glicol_FTIR':     (1, 5),
    'Fuligem':         (1.0, 2.0),
}

# --- CLEANUP DE ARQUIVOS TEMPORÁRIOS ÓRFÃOS (PREVINE ERRO DE DISCO CHEIO) ---
try:
    _tmp_dir = tempfile.gettempdir()
    _agora = time.time()
    for _ext in ["*.png", "*.pdf"]:
        for _f in glob.glob(os.path.join(_tmp_dir, _ext)):
            try:
                _basename = os.path.basename(_f)
                if (_basename.startswith("tmp") or _basename.startswith("cedro_oleo_")) and (
                        _agora - os.path.getmtime(_f)) > 1800:
                    os.remove(_f)
            except:
                pass
except:
    pass

# --- BLINDAGEM E IMPORTAÇÃO DO BANCO DE DADOS ---
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

try:
    from database import get_db_connection
    from utils_ui import load_custom_css, ui_header, ui_kpi_card, ui_empty_state
    from utils_icons import get_icon
except ImportError:
    def load_custom_css():
        pass


    def ui_header(title, subtitle, icon):
        st.title(f"{icon} {title}"); st.caption(subtitle)


    def ui_empty_state(msg, icon):
        st.info(f"{icon} {msg}")


    def get_icon(name, color, size="36"):
        return "🧪"


    def ui_kpi_card(col, title, value, icon, color, desc):
        with col:
            st.markdown(f"""
            <div style="background:white; padding:15px; border-radius:10px; border-left: 5px solid {color}; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <div style="font-size:12px; color:#64748B;">{title}</div>
                <div style="font-size:24px; font-weight:bold; color:#1E293B;">{value}</div>
                <div style="font-size:11px; color:#94A3B8;">{desc}</div>
            </div>
            """, unsafe_allow_html=True)


    def get_db_connection():
        import sqlite3
        return sqlite3.connect('manutencao.db', check_same_thread=False)

# Tentativa segura de importar pacotes para Gráficos
try:
    import matplotlib

    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.ticker as ticker
    import matplotlib.dates as mdates

    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# --- CONFIGURAÇÃO INICIAL E BANCO DE DADOS ---
load_custom_css()

icon_oleo = get_icon("droplet", "#3B82F6", "36")
ui_header(
    title="Análise Preditiva de Óleo",
    subtitle="Diagnósticos Sugeridos, Benchmarking de Desgaste e Geração Automática de O.S.",
    icon=icon_oleo
)


@contextmanager
def db_connection():
    """
    Context manager seguro para conexões SQLite.
    Garante commit em caso de sucesso, rollback em exceção,
    e sempre fecha a conexão no finally — evitando leaks.
    """
    conn = get_db_connection()
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def inicializar_tabela_feedback():
    try:
        with db_connection() as conn:
            conn.cursor().execute("""
                CREATE TABLE IF NOT EXISTS analises_oleo_feedback (
                    amostra TEXT PRIMARY KEY,
                    acao_gestao TEXT,
                    status_acao TEXT DEFAULT 'Pendente'
                )
            """)
    except Exception as e:
        st.error(f"Erro ao iniciar banco de dados de óleo: {e}")


inicializar_tabela_feedback()


def sincronizar_amostras_bd(df):
    amostras_validas = [
        (str(a).strip(),)
        for a in df['NUM_AMOSTRA'].dropna().unique()
        if str(a).strip() and str(a).strip() != '-'
    ]

    with db_connection() as conn:
        # executemany: um único round-trip no lugar de N INSERTs em loop
        conn.cursor().executemany(
            "INSERT OR IGNORE INTO analises_oleo_feedback "
            "(amostra, acao_gestao, status_acao) VALUES (?, '', 'Pendente')",
            amostras_validas
        )
        df_bd = pd.read_sql(
            "SELECT amostra as NUM_AMOSTRA, acao_gestao as ACAO_GESTAO, "
            "status_acao as STATUS_ACAO FROM analises_oleo_feedback",
            conn
        )

    df['NUM_AMOSTRA'] = df['NUM_AMOSTRA'].astype(str).str.strip()
    df_bd['NUM_AMOSTRA'] = df_bd['NUM_AMOSTRA'].astype(str).str.strip()

    cols_bd = ['ACAO_GESTAO', 'STATUS_ACAO']
    df = df.drop(columns=[c for c in cols_bd if c in df.columns])
    df = pd.merge(df, df_bd, on='NUM_AMOSTRA', how='left')
    df['ACAO_GESTAO'] = df['ACAO_GESTAO'].fillna('')
    df['STATUS_ACAO'] = df['STATUS_ACAO'].fillna('Pendente')

    return df


# ==============================================================================
# MOTOR DE PROCESSAMENTO DO LAUDO (TRADUTOR UNIVERSAL BLINDADO)
# ==============================================================================
def clean_col_name(c):
    c = str(c).lower().replace(' ', '')
    c = unicodedata.normalize('NFKD', c).encode('ASCII', 'ignore').decode('utf-8')
    c = ''.join(e for e in c if e.isalnum())
    return c


def safe_float_lab(val):
    if pd.isna(val) or val == '': return 0.0
    val_str = str(val).strip().replace('<', '').replace('>', '').replace('%', '').strip()
    if ',' in val_str: val_str = val_str.replace('.', '').replace(',', '.')
    try:
        return float(val_str)
    except:
        return 0.0


def _ler_bytes_upload(file):
    """Lê o conteúdo do arquivo e retorna (bytes, nome).
    Deve ser chamada ANTES de processar_laudo_oleo para garantir hash estável."""
    conteudo = file.read()
    file.seek(0)
    return conteudo, file.name


@st.cache_data(show_spinner="A analisar química, a descodificar e a sincronizar com a Base de Dados...", ttl=600)
def processar_laudo_oleo(file_bytes: bytes, file_name: str):
    """
    Recebe BYTES (não o file object) para garantir hash estável no cache.
    UploadedFile não é hasheável de forma confiável; bytes sim.
    """
    try:
        buffer = io.BytesIO(file_bytes)
        if file_name.lower().endswith('.csv'):
            try:
                df = pd.read_csv(buffer, sep=';', encoding='utf-8')
            except Exception:
                buffer.seek(0)
                try:
                    df = pd.read_csv(buffer, sep=';', encoding='latin-1')
                except Exception:
                    buffer.seek(0)
                    df = pd.read_csv(buffer, sep=None, engine='python', encoding='latin-1')
        else:
            df = pd.read_excel(buffer)

        original_cols = list(df.columns)
        clean_cols = [clean_col_name(c) for c in original_cols]
        available_cols = dict(zip(clean_cols, original_cols))

        def find_col(keywords, prefer_with_data=False):
            """Encontra a primeira coluna cujo nome limpo contém alguma das keywords.
            Se prefer_with_data=True (usado p/ elementos químicos), prefere a variante
            que tem dados não-zerados — necessário para laudos ALS que repetem nomes
            (Cobre, Cobre.1, Cobre.2…) onde apenas uma variante está preenchida."""
            candidatos = []
            for clean_c, orig_c in list(available_cols.items()):
                for kw in keywords:
                    if kw in clean_c:
                        candidatos.append((clean_c, orig_c))
                        break

            if not candidatos:
                return None

            escolhida_clean, escolhida_orig = candidatos[0]
            if prefer_with_data and len(candidatos) > 1:
                # Escolhe a variante com mais valores não-zerados/não-nulos
                melhor_score = -1
                for clean_c, orig_c in candidatos:
                    serie = df[orig_c]
                    # Conta valores não-nulos que parecem ter conteúdo real
                    score = serie.notna().sum()
                    if pd.api.types.is_numeric_dtype(serie):
                        score = (serie.fillna(0) != 0).sum()
                    else:
                        # texto: conta strings que parecem numéricos (contêm dígito)
                        score = serie.astype(str).str.contains(r'\d', na=False).sum()
                    if score > melhor_score:
                        melhor_score = score
                        escolhida_clean = clean_c
                        escolhida_orig = orig_c

            del available_cols[escolhida_clean]
            return escolhida_orig

        # ==================================================================
        # MAPEAMENTO DE COLUNAS — compatível com ALS e outros laboratórios
        # ==================================================================
        col_map = {
            # Identificação
            'FROTA':           find_col(['tagfrota', 'tag', 'frota', 'chassi']),
            'SERIE_CHASSI':    find_col(['seriechassi', 'serie', 'chassi', 'numeroserie']),
            'FAMILIA':         find_col(['familiadoequipamento', 'familia', 'famalia']),
            'FABRICANTE':      find_col(['fabricantedoequipamento', 'fabricante', 'marca']),
            'MODELO':          find_col(['modelodoequipamento', 'modelo']),
            'CLIENTE':         find_col(['nomedocliente', 'cliente', 'empresa']),
            'OBRA':            find_col(['obraunidade', 'obra']),
            'AREA':            find_col(['area']),
            'SETOR':           find_col(['setor']),
            # Amostra
            'NUM_AMOSTRA':     find_col(['numerodaamostra']),     # estrito: só esta coluna
            'CODIGO_LAB':      find_col(['codigolaboratorio']),    # código interno do laboratório
            'DATA_COLETA':     find_col(['datadecoleta', 'datacoleta', 'coleta']),
            'DATA_RECEBIMENTO':find_col(['dataderecebimento', 'recebimento']),
            'DATA_ENTRADA_LAB':find_col(['datadeentradanolaboratorio', 'entradanolaboratorio', 'entradaem']),
            'DATA_LIBERACAO':  find_col(['datadeliberacaodoresultado', 'liberacao', 'resultado']),
            # Status e parecer
            'STATUS_LAUDO':    find_col(['statusdaamostra', 'status', 'condicao']),
            'AVALIACAO':       find_col(['avaliacao', 'avaliao', 'parecer']),
            'RESPONSAVEL_ALS': find_col(['responsavelpelaavaliacacao', 'responsavel', 'analistals', 'analista']),
            'ACOES_INSPECAO':  find_col(['acoesdeinspecao', 'inspeao', 'inspecao']),
            'COMENTARIO_COLETA': find_col(['comentariodacoleta', 'comentario']),
            'PLANO_ANALISE':   find_col(['planodeanalise', 'planoanal', 'plano']),
            # Compartimento
            'COMPARTIMENTO':   find_col(['nomedocompartimento', 'compartimento', 'componente']),
            'TIPO_COMP_ALS':   find_col(['tipodecompartimento', 'tipocompartimento']),
            # Equipamento / óleo
            'HORAS_EQUIP':     find_col(['horasdoequipamentonacoleta', 'horasdoequipamento', 'hequip']),
            'HORAS_OLEO':      find_col(['horasdooleo', 'holeo']),
            'OLEO_TROCADO':    find_col(['oleotrocado', 'trocaoleo', 'trocado']),
            'FABRICANTE_OLEO': find_col(['fabricantedooleo', 'fabricanteoleo', 'marcaoleo']),
            'VISCOSIDADE_OLEO':find_col(['viscosidadedooleo', 'grausae', 'sae']),
            'INDICE_PQ':       find_col(['indicepq', 'pqindex', 'pqindex']),
        }

        if not col_map['FROTA'] or not col_map['STATUS_LAUDO']:
            st.error("Não foi possível encontrar as colunas de 'Tag / Frota' ou 'Status' no seu ficheiro.")
            return None

        rename_dict = {v: k for k, v in col_map.items() if v is not None}
        df = df.rename(columns=rename_dict)

        # Limpeza de strings
        for c in ['FROTA', 'NUM_AMOSTRA']:
            if c in df.columns:
                df[c] = df[c].astype(str).str.replace(r'\.0$', '', regex=True).str.replace('nan', '-')

        # Datas
        for c in ['DATA_COLETA', 'DATA_LIBERACAO', 'DATA_RECEBIMENTO', 'DATA_ENTRADA_LAB']:
            if c in df.columns:
                df[c] = pd.to_datetime(df[c], dayfirst=True, errors='coerce')

        # ==================================================================
        # MAPEAMENTO QUÍMICO COMPLETO — todos os elementos da ALS
        # Ordem: desgaste → contaminação → aditivos → FTIR → integridade
        # NOTA: Quando a ALS exporta, muitos elementos aparecem em colunas
        # repetidas (Cobre, Cobre.1, Cobre.2…). Apenas uma é preenchida por
        # amostra — usamos prefer_with_data=True para pegar a variante com dados.
        # ==================================================================
        metais_busca = {
            # --- Metais de desgaste (estrutural) ---
            'Ferro':      ['ferro'],
            'Cromo':      ['cromo'],
            'Níquel':     ['niquel'],
            'Chumbo':     ['chumbo'],
            'Estanho':    ['estanho'],
            'Cobre':      ['cobre'],
            'Alumínio':   ['aluminio', 'alumnio'],
            'Manganês':   ['manganes'],
            'Titânio':    ['titanio'],
            'Vanádio':    ['vanadio'],
            'Prata':      ['prata'],
            'Cádmio':     ['cadmio'],
            # --- Contaminantes externos ---
            'Silício':    ['silicio', 'silcio'],
            'Sódio':      ['sodio'],
            'Potássio':   ['potassio'],
            'Boro':       ['boro'],
            # --- Aditivos / baseline do óleo ---
            'Magnésio':   ['magnesio'],
            'Cálcio':     ['calcio'],
            'Fósforo':    ['fosforo'],
            'Zinco':      ['zinco'],
            'Bário':      ['bario'],
            'Molibdênio': ['molibdenio'],
            # --- Condição do óleo (keywords específicas, evitam conflito) ---
            'Viscosidade':     ['viscosidade100c', 'viscosidadea100', 'viscosidade100', 'v100', 'cst'],
            'Diluição Diesel': ['diluicaopordiesel', 'diluicao'],
            'TBN':             ['tbn4739', 'tbn2896', 'tbn', 'numerobas', 'basico'],
            'KF_Agua':         ['kfvolumetrico', 'kfcoulometrico', 'karlfischer'],
            # --- FTIR (sufixo 'ftir' obrigatório para distinguir das contagens visuais) ---
            'Fuligem':         ['fuligemftir', 'absorbanciamostra'],
            'Oxidação_FTIR':   ['oxidacaoftir', 'oxidacao'],
            'Nitração_FTIR':   ['nitracaoftir', 'nitracao'],
            'Sulfatação_FTIR': ['sulfatacaoftir', 'sulfatacao'],
            'Glicol_FTIR':     ['glicolftir', 'glicol'],
            'Agua_FTIR':       ['ftiraguaoleo', 'ftiragua'],
        }

        colunas_quimicas_encontradas = []
        for nome_sistema, palavras_chave in metais_busca.items():
            col_match = find_col(palavras_chave, prefer_with_data=True)
            if col_match and col_match in df.columns:
                df[nome_sistema] = df[col_match].apply(safe_float_lab)
                colunas_quimicas_encontradas.append(nome_sistema)
            else:
                df[nome_sistema] = 0.0

        # INDICE_PQ separado (busca específica para PQ Index)
        if 'INDICE_PQ' in df.columns:
            df['INDICE_PQ'] = df['INDICE_PQ'].apply(safe_float_lab)
        else:
            col_pq = find_col(['pqindex', 'indicepq', 'pq'])
            df['INDICE_PQ'] = df[col_pq].apply(safe_float_lab) if col_pq and col_pq in df.columns else 0.0
        if df['INDICE_PQ'].sum() > 0:
            colunas_quimicas_encontradas.append('INDICE_PQ')

        # Flag de troca de óleo
        if 'OLEO_TROCADO' in df.columns:
            df['OLEO_TROCADO'] = df['OLEO_TROCADO'].astype(str).str.upper()
            df['TROCOU_OLEO_FLAG'] = df['OLEO_TROCADO'].apply(lambda x: 'Sim' if x.startswith('S') else 'Não')
        else:
            df['TROCOU_OLEO_FLAG'] = 'Não'

        # ==================================================================
        # CAPTURA DOS FLAGS DE STATUS POR ELEMENTO (exclusivo ALS)
        # A ALS sinaliza quais elementos específicos estão fora do limite.
        # Essas colunas aparecem com sufixos .1/.2/.3 no Excel exportado.
        # ==================================================================
        ELEMENTOS_STATUS_ALS = ['Ferro', 'Silício', 'Cromo', 'Níquel', 'Potássio',
                                 'Sódio', 'Alumínio', 'Cobre', 'Chumbo', 'Viscosidade']
        df['ELEMENTOS_EM_ALERTA_ALS'] = ''
        for elem in ELEMENTOS_STATUS_ALS:
            # Procura colunas com sufixo duplicado que sejam status (contêm Anormal/Atenção/Crítico)
            for col in df.columns:
                col_clean = clean_col_name(col)
                elem_clean = clean_col_name(elem)
                if col_clean.startswith(elem_clean) and col != elem and col not in rename_dict.values():
                    uniq = df[col].dropna().astype(str).str.upper().unique()
                    if any(x in ' '.join(uniq) for x in ['ANORMAL', 'ATENCAO', 'CRITICO']):
                        # É um flag de status deste elemento
                        mask = df[col].notna() & (df[col].astype(str).str.upper() != 'NORMAL')
                        df.loc[mask, 'ELEMENTOS_EM_ALERTA_ALS'] = (
                            df.loc[mask, 'ELEMENTOS_EM_ALERTA_ALS'] + elem + ', '
                        )
                        break
        df['ELEMENTOS_EM_ALERTA_ALS'] = df['ELEMENTOS_EM_ALERTA_ALS'].str.rstrip(', ')

        # ==================================================================
        # DADOS RELEVANTES — resumo legível para exibição
        # ==================================================================
        COLS_RESUMO = ['Ferro', 'Silício', 'Cromo', 'Níquel', 'Cobre', 'Alumínio',
                       'Sódio', 'Potássio', 'Manganês', 'Chumbo', 'Estanho', 'Molibdênio',
                       'Fuligem', 'INDICE_PQ', 'TBN', 'Diluição Diesel', 'Viscosidade',
                       'Oxidação_FTIR', 'Nitração_FTIR', 'Sulfatação_FTIR', 'Glicol_FTIR']

        def compilar_relevantes(row):
            relevantes = []
            for col in COLS_RESUMO:
                if col in row.index:
                    val = row.get(col, 0)
                    if pd.notna(val) and val > 0:
                        relevantes.append(f"{col}: {val:g}")
            return " | ".join(relevantes) if relevantes else "-"

        df['DADOS_RELEVANTES'] = df.apply(compilar_relevantes, axis=1)

        # ==================================================================
        # BUG 2 — col_or_default garante Série mesmo quando coluna não existe
        # ==================================================================
        def col_or_default(coluna, default):
            if coluna in df.columns:
                return df[coluna].fillna(default)
            return pd.Series([default] * len(df), index=df.index)

        df['STATUS_LAUDO'] = col_or_default('STATUS_LAUDO', 'Normal')

        def padronizar_status(val):
            v = str(val).upper()
            if any(x in v for x in ['CRÍT', 'CRIT', 'INTERVENÇÃO', 'AÇÃO', 'VERMELHO']): return 'Crítico'
            if any(x in v for x in ['ALERT', 'ATENÇÃO', 'MONITORAR', 'AMARELO', 'ANORMAL', 'LARANJA']): return 'Alerta'
            return 'Normal'

        df['STATUS_CORRIGIDO'] = df['STATUS_LAUDO'].apply(padronizar_status)

        # Metadados completos
        df['AVALIACAO']          = col_or_default('AVALIACAO',           'Sem avaliação do lab.')
        df['ACOES_INSPECAO']     = col_or_default('ACOES_INSPECAO',      'Nenhuma ação sugerida.')
        df['RESPONSAVEL_ALS']    = col_or_default('RESPONSAVEL_ALS',     '-')
        df['PLANO_ANALISE']      = col_or_default('PLANO_ANALISE',       '-')
        df['COMENTARIO_COLETA']  = col_or_default('COMENTARIO_COLETA',   '-')
        df['Família do equipamento'] = col_or_default('FAMILIA',         'N/A')
        df['FABRICANTE']         = col_or_default('FABRICANTE',          '-')
        df['MODELO']             = col_or_default('MODELO',              '-')
        df['SERIE_CHASSI']       = col_or_default('SERIE_CHASSI',        '-')
        df['CLIENTE']            = col_or_default('CLIENTE',             'N/I')
        df['OBRA']               = col_or_default('OBRA',                'N/I')
        df['AREA']               = col_or_default('AREA',                '-')
        df['FABRICANTE_OLEO']    = col_or_default('FABRICANTE_OLEO',     '-')
        df['VISCOSIDADE_OLEO']   = col_or_default('VISCOSIDADE_OLEO',    '-')
        df['HORAS_OLEO']         = col_or_default('HORAS_OLEO',          0)
        df['HORAS_EQUIP']        = col_or_default('HORAS_EQUIP',         0)

        # ==================================================================
        # TIPO DE COMPARTIMENTO — usa classificação da ALS quando disponível
        # ==================================================================
        def classificar_compartimento(val):
            v = str(val).upper()
            if 'MOTOR' in v: return 'Motor'
            if any(x in v for x in ['TRANSMISS', 'CAIXA', 'CÂMBIO', 'CONVERSOR']): return 'Transmissão'
            if 'DIFERENCIAL' in v or 'EIXO' in v: return 'Diferencial'
            if 'CUBO' in v or 'COMANDO FINAL' in v or 'RODA' in v: return 'Cubos/Comandos Finais'
            if 'HIDRÁULIC' in v or 'HIDRAULIC' in v: return 'Hidráulico'
            if 'REDUT' in v or 'REDUTOR' in v: return 'Redutor'
            return 'Outros'

        # Prioriza o tipo já classificado pela ALS; usa compartimento como fallback
        tipo_base = col_or_default('TIPO_COMP_ALS', '').where(
            col_or_default('TIPO_COMP_ALS', '') != '', col_or_default('COMPARTIMENTO', 'Outros')
        )
        df['TIPO_COMPARTIMENTO'] = tipo_base.apply(classificar_compartimento)

        # ==================================================================
        # DIAGNÓSTICO IA — usa flags de elemento da ALS + análise própria
        # ==================================================================
        def gerar_diagnostico_ia(r):
            if r['STATUS_CORRIGIDO'] == 'Normal':
                return "✅ Sistema a operar dentro dos parâmetros."

            alertas = []

            # 1. Usa os flags por elemento da ALS (mais confiáveis que thresholds fixos)
            elementos_als = str(r.get('ELEMENTOS_EM_ALERTA_ALS', ''))
            if elementos_als and elementos_als != 'nan':
                for elem in [e.strip() for e in elementos_als.split(',') if e.strip()]:
                    if elem == 'Ferro':
                        alertas.append("🔧 Ferro elevado: desgaste anormal de componentes internos.")
                    elif elem == 'Silício':
                        alertas.append("🌪️ Silício elevado: contaminação por areia/poeira (filtro de ar).")
                    elif elem in ['Cromo', 'Níquel']:
                        alertas.append(f"⚙️ {elem} elevado: desgaste em anéis, camisas ou eixos.")
                    elif elem == 'Alumínio':
                        alertas.append("🔩 Alumínio elevado: desgaste em pistões ou mancais de alumínio.")
                    elif elem in ['Sódio', 'Potássio']:
                        alertas.append("💧 Sódio/Potássio elevado: contaminação por líquido de arrefecimento.")
                    elif elem == 'Cobre':
                        alertas.append("🟤 Cobre elevado: desgaste em bronzinas ou trocadores de calor.")

            # 2. Análise complementar com dados químicos (captura o que a ALS não sinaliza por elemento)
            fe  = r.get('Ferro',       0)
            si  = r.get('Silício',     0)
            cu  = r.get('Cobre',       0)
            na  = r.get('Sódio',       0)
            k   = r.get('Potássio',    0)
            ni  = r.get('Níquel',      0)
            cr  = r.get('Cromo',       0)
            pb  = r.get('Chumbo',      0)
            sn  = r.get('Estanho',     0)
            mn  = r.get('Manganês',    0)
            pq  = r.get('INDICE_PQ',   0)
            dd  = r.get('Diluição Diesel', 0)
            ful = r.get('Fuligem',     0)
            tbn = r.get('TBN',         0)
            oxi = r.get('Oxidação_FTIR', 0)
            nit = r.get('Nitração_FTIR', 0)
            sul = r.get('Sulfatação_FTIR', 0)
            gli = r.get('Glicol_FTIR', 0)
            kf  = r.get('KF_Agua',     0)

            # Só adiciona se não foi coberto pelos flags ALS
            ja_cobertos = alertas

            if si > 15 and fe > 15 and not any('Silício' in a or 'poeira' in a for a in ja_cobertos):
                alertas.append("🌪️ Silício + Ferro: entrada de poeira acelerando desgaste.")
            if pb > 10 and not any('Chumbo' in a for a in ja_cobertos):
                alertas.append("⚙️ Chumbo elevado: desgaste em bronzinas de mancal.")
            if sn > 10:
                alertas.append("🔩 Estanho elevado: desgaste em buchas ou rolamentos de bronze.")
            if mn > 5:
                alertas.append("⚙️ Manganês elevado: desgaste em componentes de aço-manganês.")
            if pq > 40:
                alertas.append("🧲 PQ Index alto: presença de partículas ferrosas grandes — possível fadiga.")
            if dd > 4.0:
                alertas.append("⛽ Diluição por diesel > 4%: combustível no óleo.")
            if ful > 1.5:
                alertas.append("⬛ Fuligem (FTIR) alta: falha de queima, filtros ou injetores.")
            if tbn > 0 and tbn < 5.0:
                alertas.append("⚗️ TBN baixo (< 5): óleo com neutralização esgotada — trocar.")
            if oxi > 20:
                alertas.append("🔥 Oxidação (FTIR) elevada: óleo degradado termicamente.")
            if nit > 20:
                alertas.append("💨 Nitração (FTIR) elevada: blow-by ou combustão incompleta.")
            if sul > 30:
                alertas.append("🧪 Sulfatação (FTIR) elevada: combustível com enxofre no cárter.")
            if gli > 5:
                alertas.append("💧 Glicol (FTIR) detectado: contaminação por aditivo de arrefecimento.")
            if kf > 500:
                alertas.append("💦 Água (KF) elevada: umidade no sistema — risco de corrosão.")

            if not alertas:
                return "⚠️ Verificar laudo original — elemento(s) acima do limite não identificados aqui."
            return "\n".join(alertas)

        df['DIAGNOSTICO_IA'] = df.apply(gerar_diagnostico_ia, axis=1)

        # ==================================================================
        # HEALTH SCORE (0-100) — agregação ponderada da química
        # ==================================================================
        # Pesos por categoria (somam ~100 quando todos no limite)
        # Metais de desgaste pesam mais (60), contaminantes (25), condição (15)
        PESOS_HEALTH = {
            # Metais críticos: cada um pode descontar até X pontos
            'Ferro':      ('desgaste', 15),
            'Cromo':      ('desgaste', 10),
            'Níquel':     ('desgaste', 8),
            'Cobre':      ('desgaste', 7),
            'Chumbo':     ('desgaste', 7),
            'Alumínio':   ('desgaste', 6),
            'Estanho':    ('desgaste', 4),
            'Manganês':   ('desgaste', 3),
            # Contaminantes
            'Silício':    ('contam',   12),
            'Sódio':      ('contam',   7),
            'Potássio':   ('contam',   6),
            # Condição
            'TBN':        ('cond',     6),   # inverso
            'Diluição Diesel': ('cond', 5),
            'KF_Agua':    ('cond',     4),
        }

        def calcular_health_score(r):
            """
            Retorna score 0-100 (100 = perfeito, 0 = severo).

            REGRA PRIMÁRIA: a classificação da ALS é a referência final.
            - Se ALS = Crítico → score ≤ 35 (sempre faixa Crítico)
            - Se ALS = Alerta/Anormal → score ≤ 55 (sempre faixa Atenção ou pior)
            - Se ALS = Normal → score técnico (livre, 0-100)

            Isso garante que o gráfico de distribuição reflete o veredito da ALS,
            mas o score técnico (química) ainda diferencia amostras Normal entre si.
            """
            # --- Score técnico baseado em química ---
            penalidade_total = 0.0
            for elem, (cat, peso) in PESOS_HEALTH.items():
                if elem not in r.index or elem not in LIMITES_REFERENCIA:
                    continue
                val = r.get(elem, 0)
                try:
                    val = float(val) if pd.notna(val) else 0.0
                except (ValueError, TypeError):
                    val = 0.0
                lv, la = LIMITES_REFERENCIA[elem]
                # Caso TBN: lógica invertida (baixo = ruim)
                if elem == 'TBN':
                    if val == 0:
                        continue  # sem dado, não penaliza
                    if val >= lv:
                        ratio = 0
                    elif val >= la:
                        ratio = (lv - val) / (lv - la) * 0.5  # alerta
                    else:
                        ratio = 0.5 + min((la - val) / la, 1.0) * 0.5  # crítico
                else:
                    if val <= lv:
                        ratio = 0
                    elif val <= la:
                        ratio = (val - lv) / (la - lv) * 0.5  # 0-50% peso
                    else:
                        # Acima do amarelo: penaliza forte, max 1.0 (peso total)
                        ratio = 0.5 + min((val - la) / la, 1.0) * 0.5
                penalidade_total += ratio * peso

            score_tecnico = max(0, min(100, 100 - penalidade_total))

            # --- Aplica teto pelo veredito da ALS (alinhamento com o laboratório) ---
            status = r.get('STATUS_CORRIGIDO', 'Normal')
            if status == 'Crítico':
                score_final = min(score_tecnico, 35.0)
            elif status == 'Alerta':
                score_final = min(score_tecnico, 55.0)
            else:
                score_final = score_tecnico

            return round(score_final, 1)

        df['HEALTH_SCORE'] = df.apply(calcular_health_score, axis=1)

        # ==================================================================
        # TENDÊNCIA — taxa de crescimento entre as últimas 2 amostras
        # da mesma (FROTA, COMPARTIMENTO)
        # ==================================================================
        df_sorted = df.sort_values(['FROTA', 'COMPARTIMENTO', 'DATA_COLETA'])
        ELEMENTOS_TENDENCIA = ['Ferro', 'Silício', 'Cromo', 'Níquel', 'Cobre',
                                'Alumínio', 'Sódio', 'Potássio']

        def calcular_tendencia(grupo):
            """Retorna 'estabilizado'|'crescimento normal'|'acelerado'|'explosivo'|'sem dados'."""
            if len(grupo) < 2:
                grupo['TENDENCIA'] = 'sem histórico'
                grupo['MAX_CRESCIMENTO_PCT'] = 0.0
                return grupo

            ultima = grupo.iloc[-1]
            anterior = grupo.iloc[-2]

            max_pct = 0.0
            for elem in ELEMENTOS_TENDENCIA:
                if elem not in grupo.columns:
                    continue
                v_ant = float(anterior.get(elem, 0) or 0)
                v_ult = float(ultima.get(elem, 0) or 0)
                # Só considera se valor atual > 5 ppm (evita ruído de zeros)
                if v_ult < 5:
                    continue
                if v_ant > 0:
                    pct = ((v_ult - v_ant) / v_ant) * 100
                elif v_ult > 10:
                    pct = 999  # disparou de zero
                else:
                    pct = 0
                if pct > max_pct:
                    max_pct = pct

            # Classifica
            if max_pct < 10:
                tend = 'estabilizado'
            elif max_pct < 50:
                tend = 'crescimento normal'
            elif max_pct < 150:
                tend = 'acelerado'
            else:
                tend = 'explosivo'

            grupo['TENDENCIA'] = 'sem histórico'  # default das amostras anteriores
            grupo['MAX_CRESCIMENTO_PCT'] = 0.0
            # Aplica só na última amostra
            grupo.iloc[-1, grupo.columns.get_loc('TENDENCIA')] = tend
            grupo.iloc[-1, grupo.columns.get_loc('MAX_CRESCIMENTO_PCT')] = round(max_pct, 1)
            return grupo

        df['TENDENCIA'] = 'sem histórico'
        df['MAX_CRESCIMENTO_PCT'] = 0.0
        try:
            df = df_sorted.groupby(['FROTA', 'COMPARTIMENTO'], group_keys=False).apply(calcular_tendencia)
        except Exception as e:
            logger.warning("Cálculo de tendência falhou (usando defaults): %s", e)

        # ==================================================================
        # RPN — Risk Priority Number
        # RPN = (100 - HealthScore) * mult_tendencia * mult_compartimento
        # ==================================================================
        MULT_TENDENCIA = {
            'estabilizado':       1.0,
            'crescimento normal': 1.2,
            'acelerado':          1.5,
            'explosivo':          2.0,
            'sem histórico':      1.1,  # incerteza
        }
        MULT_COMPARTIMENTO = {
            'Motor':              1.3,
            'Transmissão':        1.2,
            'Diferencial':        1.1,
            'Cubos/Comandos Finais': 1.0,
            'Hidráulico':         1.0,
            'Redutor':            1.0,
            'Outros':             1.0,
        }

        def calcular_rpn(r):
            base = 100 - r.get('HEALTH_SCORE', 100)
            mt = MULT_TENDENCIA.get(r.get('TENDENCIA', 'sem histórico'), 1.1)
            mc = MULT_COMPARTIMENTO.get(r.get('TIPO_COMPARTIMENTO', 'Outros'), 1.0)
            return round(base * mt * mc, 1)

        df['RPN'] = df.apply(calcular_rpn, axis=1)

        df = sincronizar_amostras_bd(df)
        return df
    except Exception as e:
        st.error(f"Erro na análise de dados: {e}")
        return None


# ==============================================================================
# GERAÇÃO DE EXCEL E PDF
# ==============================================================================

@st.cache_data(show_spinner="A gerar Planilha de Ações (Excel)...", ttl=60)
def gerar_excel_plano_acao(df_export):
    if not OPENPYXL_AVAILABLE: return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Plano de Ação - Óleo"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    action_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    action_font = Font(color="A16207", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    colunas = [
        "Amostra", "Frota", "Família", "Fabricante", "Modelo", "Compartimento", "Data Coleta",
        "Horas Equip.", "Horas Óleo", "Trocou Óleo",
        "Status ALS", "Health Score", "RPN", "Tendência",
        "Avaliação do Laboratório", "Ações de Inspeção Recomendadas",
        "Analista ALS", "Plano de Análise", "Diagnóstico Sugerido", "Elementos sinalizados ALS",
        # Metais de desgaste
        "Fe", "Cr", "Ni", "Pb", "Sn", "Cu", "Al", "Mn", "Ti", "V", "Ag", "Cd",
        # Contaminantes
        "Si", "Na", "K", "B",
        # Aditivos
        "Mg", "Ca", "P", "Zn", "Ba", "Mo",
        # Condição
        "Visc. 100°C", "TBN", "Diluição Diesel %", "Água (KF)", "PQ Index",
        # FTIR
        "Fuligem", "Oxidação", "Nitração", "Sulfatação", "Glicol", "FTIR-Água",
        # Ação
        "AÇÃO DA GESTÃO / RETORNO"
    ]

    for col_num, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        if "RETORNO" in col_name:
            cell.fill = action_fill;
            cell.font = action_font
        else:
            cell.fill = header_fill;
            cell.font = header_font
        cell.alignment = align_center;
        cell.border = border

    # Mapeamento ordenado (label → chave no df, casas decimais)
    mapping_data = [
        # Identificação
        (lambda r: str(r.get('NUM_AMOSTRA', '')), align_center, None),
        (lambda r: str(r.get('FROTA', '')), align_center, None),
        (lambda r: str(r.get('Família do equipamento', '')), align_center, None),
        (lambda r: str(r.get('FABRICANTE', '')), align_center, None),
        (lambda r: str(r.get('MODELO', '')), align_center, None),
        (lambda r: str(r.get('COMPARTIMENTO', '')), align_center, None),
        (lambda r: r['DATA_COLETA'].strftime('%d/%m/%Y') if pd.notnull(r.get('DATA_COLETA')) else '-', align_center, None),
        (lambda r: r.get('HORAS_EQUIP', 0), align_center, None),
        (lambda r: r.get('HORAS_OLEO', 0), align_center, None),
        (lambda r: str(r.get('TROCOU_OLEO_FLAG', 'Não')), align_center, None),
        # Status
        (lambda r: str(r.get('STATUS_CORRIGIDO', '')), align_center, 'status'),
        (lambda r: r.get('HEALTH_SCORE', 0), align_center, 'health'),
        (lambda r: r.get('RPN', 0), align_center, 'rpn'),
        (lambda r: str(r.get('TENDENCIA', '-')), align_center, None),
        (lambda r: str(r.get('AVALIACAO', '')), align_left, None),
        (lambda r: str(r.get('ACOES_INSPECAO', '')), align_left, None),
        (lambda r: str(r.get('RESPONSAVEL_ALS', '-')), align_center, None),
        (lambda r: str(r.get('PLANO_ANALISE', '-')), align_center, None),
        (lambda r: str(r.get('DIAGNOSTICO_IA', '')), align_left, None),
        (lambda r: str(r.get('ELEMENTOS_EM_ALERTA_ALS', '')), align_center, None),
    ]
    # Elementos químicos
    elementos_excel = [
        'Ferro', 'Cromo', 'Níquel', 'Chumbo', 'Estanho', 'Cobre',
        'Alumínio', 'Manganês', 'Titânio', 'Vanádio', 'Prata', 'Cádmio',
        'Silício', 'Sódio', 'Potássio', 'Boro',
        'Magnésio', 'Cálcio', 'Fósforo', 'Zinco', 'Bário', 'Molibdênio',
        'Viscosidade', 'TBN', 'Diluição Diesel', 'KF_Agua', 'INDICE_PQ',
        'Fuligem', 'Oxidação_FTIR', 'Nitração_FTIR', 'Sulfatação_FTIR', 'Glicol_FTIR', 'Agua_FTIR',
    ]
    for elem in elementos_excel:
        mapping_data.append((lambda r, e=elem: r.get(e, 0), align_center, None))

    # Ação (última coluna, especial)
    for row_num, (_, r) in enumerate(df_export.iterrows(), 2):
        for col_idx, (extractor, alignment, special) in enumerate(mapping_data, 1):
            valor = extractor(r)
            cell = ws.cell(row=row_num, column=col_idx, value=valor)
            cell.alignment = alignment
            if special == 'status':
                if r.get('STATUS_CORRIGIDO') == 'Crítico':
                    cell.font = Font(color="DC2626", bold=True)
                elif r.get('STATUS_CORRIGIDO') == 'Alerta':
                    cell.font = Font(color="D97706", bold=True)
                else:
                    cell.font = Font(color="16A34A", bold=True)
            elif special == 'health':
                hs = r.get('HEALTH_SCORE', 0)
                if hs >= 75:
                    cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                    cell.font = Font(color="166534", bold=True)
                elif hs >= 55:
                    cell.fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
                    cell.font = Font(color="854D0E", bold=True)
                else:
                    cell.fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
                    cell.font = Font(color="7F1D1D", bold=True)
            elif special == 'rpn':
                rpn = r.get('RPN', 0)
                if rpn >= 80:
                    cell.fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
                    cell.font = Font(color="7F1D1D", bold=True)
                elif rpn >= 40:
                    cell.fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
                    cell.font = Font(color="854D0E", bold=True)

        # Coluna final: Ação da Gestão
        acao_bd = str(r.get('ACAO_GESTAO', '')).strip()
        action_c = ws.cell(row=row_num, column=len(colunas), value=acao_bd)
        action_c.fill = PatternFill(start_color="FEFCE8", end_color="FEFCE8", fill_type="solid")
        action_c.border = border
        action_c.alignment = align_left

    # Larguras de colunas — adaptadas
    larguras = {
        'A': 14, 'B': 12, 'C': 18, 'D': 14, 'E': 14, 'F': 22, 'G': 11,
        'H': 9, 'I': 9, 'J': 9,
        'K': 11, 'L': 45, 'M': 35, 'N': 22, 'O': 22, 'P': 45, 'Q': 25,
    }
    for letra, w in larguras.items():
        ws.column_dimensions[letra].width = w
    # Elementos químicos: 7 chars cada
    from openpyxl.utils import get_column_letter
    for i in range(18, len(colunas)):  # depois das colunas de identificação
        ws.column_dimensions[get_column_letter(i)].width = 8
    ws.column_dimensions[get_column_letter(len(colunas))].width = 55

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ==============================================================================
# HELPERS DE FORMATAÇÃO E GRÁFICOS PARA O PDF
# ==============================================================================
# Nota: LIMITES_REFERENCIA está definido no topo do arquivo (constante global)


def _safe_latin(val):
    """Helper único para a conversão repetitiva latin-1."""
    return str(val).encode('latin-1', 'ignore').decode('latin-1')


def _cor_status_elemento(elemento, valor):
    """Retorna RGB para colorir célula de um elemento baseado nos limites.
    Retorna (R, G, B, status) — status é 'normal'/'alerta'/'critico'."""
    if elemento not in LIMITES_REFERENCIA or valor is None or pd.isna(valor):
        return (255, 255, 255, 'sem_ref')
    try:
        valor = float(valor)
    except (ValueError, TypeError):
        return (255, 255, 255, 'sem_ref')

    limite_verde, limite_amarelo = LIMITES_REFERENCIA[elemento]
    # Caso especial TBN: lógica inversa (baixo = ruim)
    if elemento == 'TBN':
        if valor >= limite_verde:
            return (220, 252, 231, 'normal')      # verde claro
        if valor >= limite_amarelo:
            return (254, 240, 138, 'alerta')      # amarelo claro
        return (254, 202, 202, 'critico')         # vermelho claro

    if valor <= limite_verde:
        return (220, 252, 231, 'normal')
    if valor <= limite_amarelo:
        return (254, 240, 138, 'alerta')
    return (254, 202, 202, 'critico')


def _gerar_radar_desgaste(row, output_path):
    """
    Gera um gráfico tipo radar (spider) com os 8 metais de desgaste principais.
    Cada eixo é normalizado pelo limite amarelo do elemento (1.0 = limite).
    """
    if not MATPLOTLIB_AVAILABLE:
        return False
    try:
        elementos = ['Ferro', 'Cromo', 'Níquel', 'Chumbo',
                     'Cobre', 'Alumínio', 'Estanho', 'Manganês']
        valores_norm = []
        labels = []
        for elem in elementos:
            if elem in LIMITES_REFERENCIA:
                limite_amarelo = LIMITES_REFERENCIA[elem][1]
                v = float(row.get(elem, 0) or 0)
                # Normaliza: 1.0 = limite amarelo (atenção)
                valores_norm.append(min(v / limite_amarelo, 2.5))  # cap em 2.5x
                labels.append(elem)

        if not any(valores_norm):
            return False

        # Fecha o polígono
        angulos = np.linspace(0, 2 * np.pi, len(labels), endpoint=False).tolist()
        valores_norm += valores_norm[:1]
        angulos += angulos[:1]

        fig, ax = plt.subplots(figsize=(4.5, 4.5), subplot_kw=dict(polar=True))

        # Zonas de referência
        ax.fill_between(np.linspace(0, 2 * np.pi, 100), 0, 1.0,
                        color='#10B981', alpha=0.12, label='Normal')
        ax.fill_between(np.linspace(0, 2 * np.pi, 100), 1.0, 1.5,
                        color='#F59E0B', alpha=0.12, label='Alerta')
        ax.fill_between(np.linspace(0, 2 * np.pi, 100), 1.5, 2.5,
                        color='#EF4444', alpha=0.12, label='Crítico')

        # Linha do equipamento
        ax.plot(angulos, valores_norm, 'o-', linewidth=2,
                color='#1E293B', markersize=5)
        ax.fill(angulos, valores_norm, alpha=0.25, color='#1E293B')

        ax.set_xticks(angulos[:-1])
        ax.set_xticklabels(labels, fontsize=8, fontweight='bold')
        ax.set_ylim(0, 2.5)
        ax.set_yticks([0.5, 1.0, 1.5, 2.0])
        ax.set_yticklabels(['½ limite', 'limite', '1,5×', '2×'], fontsize=6, color='#666666')
        ax.grid(True, linestyle='--', alpha=0.4)
        ax.set_title('Perfil de Desgaste (normalizado vs. limite)',
                     fontsize=9, color='#1E293B', pad=15, fontweight='bold')

        plt.tight_layout()
        fig.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        return True
    except Exception as e:
        logger.warning("Radar não gerado: %s", e)
        try:
            plt.close('all')
        except Exception:
            pass
        return False


# ==============================================================================
# CAPA EXECUTIVA — Páginas iniciais do PDF para a diretoria
# ==============================================================================

def _gerar_heatmap_frota_compartimento(df_pdf, output_path, top_n=15):
    """
    Gera heatmap (frotas top_n por risco) × (compartimentos) colorido pelo Health Score.
    """
    if not MATPLOTLIB_AVAILABLE or df_pdf.empty:
        return False
    try:
        # Para cada combinação (FROTA, COMPARTIMENTO), pega a amostra MAIS RECENTE
        df_recente = (df_pdf.sort_values('DATA_COLETA')
                            .groupby(['FROTA', 'COMPARTIMENTO'])
                            .tail(1))
        # Pivot: frotas (linhas) × compartimentos (colunas)
        pivot = df_recente.pivot_table(
            index='FROTA', columns='COMPARTIMENTO',
            values='HEALTH_SCORE', aggfunc='min'
        )
        if pivot.empty:
            return False

        # Ordena frotas pela MENOR saúde (mais críticas no topo)
        pivot['_min'] = pivot.min(axis=1)
        pivot = pivot.sort_values('_min').head(top_n).drop(columns=['_min'])

        fig, ax = plt.subplots(figsize=(8, max(3, 0.35 * len(pivot) + 1)))
        # Colormap verde→amarelo→vermelho (invertido: baixo score = vermelho)
        import matplotlib.colors as mcolors
        cmap = mcolors.LinearSegmentedColormap.from_list(
            'health', ['#DC2626', '#F59E0B', '#FEF08A', '#86EFAC', '#16A34A']
        )
        im = ax.imshow(pivot.values, cmap=cmap, vmin=0, vmax=100, aspect='auto')

        ax.set_xticks(range(len(pivot.columns)))
        ax.set_xticklabels([str(c)[:18] for c in pivot.columns],
                            rotation=30, ha='right', fontsize=7)
        ax.set_yticks(range(len(pivot.index)))
        ax.set_yticklabels([str(i)[:15] for i in pivot.index], fontsize=7)

        # Anota os scores em cada célula
        for i in range(len(pivot.index)):
            for j in range(len(pivot.columns)):
                val = pivot.values[i, j]
                if not pd.isna(val):
                    cor_txt = 'white' if val < 40 or val > 75 else 'black'
                    ax.text(j, i, f'{val:.0f}', ha='center', va='center',
                            fontsize=7, color=cor_txt, fontweight='bold')

        cbar = plt.colorbar(im, ax=ax, fraction=0.025, pad=0.02)
        cbar.set_label('Health Score', fontsize=7)
        cbar.ax.tick_params(labelsize=6)

        ax.set_title(f'Mapa de Calor — Saude por Frota x Compartimento (Top {top_n} por risco)',
                      fontsize=9, fontweight='bold', pad=10, color='#1E293B')

        plt.tight_layout()
        fig.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        return True
    except Exception as e:
        logger.warning("Heatmap não gerado: %s", e)
        try: plt.close('all')
        except Exception: pass
        return False


def _gerar_pareto_elementos(df_pdf, output_path):
    """Gera Pareto dos elementos mais frequentemente sinalizados pela ALS."""
    if not MATPLOTLIB_AVAILABLE or df_pdf.empty:
        return False
    try:
        # Conta os elementos sinalizados pela ALS
        elementos = (df_pdf['ELEMENTOS_EM_ALERTA_ALS']
                       .dropna().astype(str)
                       .str.split(', ').explode().str.strip())
        elementos = elementos[elementos != ''].value_counts()
        if elementos.empty:
            return False

        elementos = elementos.head(10)
        cumulativo = (elementos.cumsum() / elementos.sum() * 100)

        fig, ax1 = plt.subplots(figsize=(6.5, 3.5))
        ax1.bar(range(len(elementos)), elementos.values, color='#DC2626', alpha=0.75)
        ax1.set_ylabel('Quantidade de amostras', fontsize=8, color='#DC2626')
        ax1.set_xticks(range(len(elementos)))
        ax1.set_xticklabels(elementos.index, rotation=30, ha='right', fontsize=7)
        ax1.tick_params(axis='y', labelsize=7, colors='#DC2626')
        ax1.set_title('Pareto — Elementos mais sinalizados',
                       fontsize=9, fontweight='bold', color='#1E293B')

        # Linha 80/20
        ax2 = ax1.twinx()
        ax2.plot(range(len(elementos)), cumulativo.values, color='#3B82F6',
                  marker='o', markersize=4, linewidth=1.5)
        ax2.axhline(y=80, color='#1E293B', linestyle='--', linewidth=0.5, alpha=0.5)
        ax2.set_ylabel('% Acumulado', fontsize=8, color='#3B82F6')
        ax2.tick_params(axis='y', labelsize=7, colors='#3B82F6')
        ax2.set_ylim(0, 105)

        for spine in ['top']:
            ax1.spines[spine].set_visible(False)
            ax2.spines[spine].set_visible(False)

        plt.tight_layout()
        fig.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        return True
    except Exception as e:
        logger.warning("Pareto não gerado: %s", e)
        try: plt.close('all')
        except Exception: pass
        return False


def _gerar_benchmark_fabricantes(df_pdf, output_path):
    """Saúde média por fabricante (barras horizontais)."""
    if not MATPLOTLIB_AVAILABLE or df_pdf.empty:
        return False
    try:
        if 'FABRICANTE' not in df_pdf.columns:
            return False
        bench = (df_pdf.groupby('FABRICANTE')
                   .agg(media=('HEALTH_SCORE', 'mean'),
                        n=('HEALTH_SCORE', 'count'))
                   .query('n >= 1')
                   .sort_values('media'))
        bench = bench[bench.index.astype(str).str.strip() != '-']
        if bench.empty:
            return False

        fig, ax = plt.subplots(figsize=(6.5, max(2, 0.3 * len(bench) + 1)))
        # Cor por faixa de saúde
        cores = ['#DC2626' if v < 40 else
                 '#F59E0B' if v < 60 else
                 '#22C55E' if v < 80 else
                 '#16A34A' for v in bench['media']]

        bars = ax.barh(range(len(bench)), bench['media'], color=cores, alpha=0.85)
        ax.set_yticks(range(len(bench)))
        ax.set_yticklabels([str(i)[:18] for i in bench.index], fontsize=7)
        ax.set_xlim(0, 100)
        ax.set_xlabel('Health Score Médio', fontsize=8)
        ax.tick_params(axis='x', labelsize=7)
        ax.set_title('Saúde média por Fabricante de equipamento',
                      fontsize=9, fontweight='bold', color='#1E293B')

        # Anotações: valor + n amostras
        for i, (_, row) in enumerate(bench.iterrows()):
            ax.text(row['media'] + 1, i,
                    f"{row['media']:.0f}  (n={int(row['n'])})",
                    va='center', fontsize=6.5, color='#1E293B')

        # Zonas verticais
        ax.axvspan(0, 40, alpha=0.04, color='#DC2626')
        ax.axvspan(40, 60, alpha=0.04, color='#F59E0B')
        ax.axvspan(60, 80, alpha=0.04, color='#FEF08A')

        for spine in ['top', 'right']:
            ax.spines[spine].set_visible(False)

        plt.tight_layout()
        fig.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        return True
    except Exception as e:
        logger.warning("Benchmark não gerado: %s", e)
        try: plt.close('all')
        except Exception: pass
        return False


def _gerar_capa_executiva(pdf, df_pdf, arquivos_temp):
    """
    Gera 2 páginas iniciais executivas:
    - Página 1: KPIs + Top 10 amostras prioritárias (por RPN)
    - Página 2: Heatmap + Pareto + Benchmark fabricantes
    """
    # ====== PÁGINA 1 — SUMÁRIO EXECUTIVO ======
    pdf.add_page()

    pdf.set_font('Arial', 'B', 14)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 8, _safe_latin('SUMÁRIO EXECUTIVO — Diretoria de Manutenção'), 0, 1, 'C')

    pdf.set_font('Arial', '', 9)
    pdf.set_text_color(100, 116, 139)
    periodo = ''
    if df_pdf['DATA_COLETA'].notna().any():
        d_min = df_pdf['DATA_COLETA'].min().strftime('%d/%m/%Y')
        d_max = df_pdf['DATA_COLETA'].max().strftime('%d/%m/%Y')
        periodo = f'Período: {d_min} a {d_max}  |  '
    pdf.cell(0, 5, _safe_latin(
        f"{periodo}{len(df_pdf)} amostras  |  {df_pdf['FROTA'].nunique()} frotas  |  "
        f"{df_pdf['COMPARTIMENTO'].nunique()} compartimentos"
    ), 0, 1, 'C')
    pdf.ln(3)

    # --- KPI STRIP (4 cards) ---
    y_kpi = pdf.get_y()
    kpi_w = 45
    kpi_h = 22

    health_medio = df_pdf['HEALTH_SCORE'].mean() if 'HEALTH_SCORE' in df_pdf.columns else 0
    n_total = len(df_pdf)
    n_crit = (df_pdf['STATUS_CORRIGIDO'] == 'Crítico').sum()
    n_alerta = (df_pdf['STATUS_CORRIGIDO'] == 'Alerta').sum()
    n_normal = (df_pdf['STATUS_CORRIGIDO'] == 'Normal').sum()

    # Percentuais
    pct_crit   = (n_crit / n_total * 100) if n_total else 0
    pct_alerta = (n_alerta / n_total * 100) if n_total else 0
    pct_normal = (n_normal / n_total * 100) if n_total else 0

    # Cor do KPI de saúde geral
    if health_medio >= 75:
        cor_saude = (22, 163, 74); rot_saude = "BOM"
    elif health_medio >= 55:
        cor_saude = (245, 158, 11); rot_saude = "ATENCAO"
    else:
        cor_saude = (220, 38, 38); rot_saude = "CRITICO"

    kpis = [
        (15,  cor_saude,        "SAUDE GERAL", f"{health_medio:.0f}/100",  rot_saude),
        (65,  (22, 163, 74),    "NORMAL",      f"{pct_normal:.0f}%",       f"{n_normal} amostras"),
        (115, (245, 158, 11),   "ANORMAL",     f"{pct_alerta:.0f}%",       f"{n_alerta} amostras"),
        (165, (220, 38, 38),    "CRITICO",     f"{pct_crit:.0f}%",         f"{n_crit} amostras"),
    ]

    for x, (r, g, b), titulo, valor, subt in kpis:
        pdf.set_fill_color(248, 250, 252)
        pdf.set_draw_color(226, 232, 240)
        pdf.rect(x, y_kpi, kpi_w, kpi_h, 'DF')
        # Faixa colorida lateral
        pdf.set_fill_color(r, g, b)
        pdf.rect(x, y_kpi, 2.5, kpi_h, 'F')

        pdf.set_xy(x + 4, y_kpi + 3)
        pdf.set_font('Arial', 'B', 7)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(40, 3, _safe_latin(titulo), 0, 1, 'L')

        pdf.set_xy(x + 4, y_kpi + 8)
        pdf.set_font('Arial', 'B', 14)
        pdf.set_text_color(r, g, b)
        pdf.cell(40, 6, _safe_latin(valor), 0, 1, 'L')

        pdf.set_xy(x + 4, y_kpi + 16)
        pdf.set_font('Arial', '', 6.5)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(40, 3, _safe_latin(subt), 0, 1, 'L')

    pdf.set_y(y_kpi + kpi_h + 6)

    # --- TOP 10 RANKING POR RPN ---
    pdf.set_font('Arial', 'B', 11)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 6, _safe_latin('Top 10 Amostras Prioritárias  (ordenadas por RPN — Risk Priority)'),
              0, 1, 'L')
    pdf.set_font('Arial', '', 8)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(0, 4, _safe_latin('RPN = (100 - Health Score) × multiplicador de tendência × multiplicador de criticidade do compartimento'),
              0, 1, 'L')
    pdf.ln(2)

    # Cabeçalho da tabela
    pdf.set_fill_color(30, 41, 59)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Arial', 'B', 7)
    larguras = [10, 28, 28, 22, 18, 16, 24, 44]
    cabecalho = ['#', 'Frota', 'Compartimento', 'Família', 'Health', 'RPN', 'Tendência', 'Ação sugerida']
    for w, h in zip(larguras, cabecalho):
        pdf.cell(w, 6, _safe_latin(h), 1, 0, 'C', fill=True)
    pdf.ln()

    # Top 10 por RPN
    top10 = df_pdf.nlargest(10, 'RPN') if 'RPN' in df_pdf.columns else df_pdf.head(10)
    pdf.set_font('Arial', '', 7)

    for i, (_, row) in enumerate(top10.iterrows(), 1):
        hs = row.get('HEALTH_SCORE', 0)
        rpn = row.get('RPN', 0)
        tend = str(row.get('TENDENCIA', '-'))[:14]

        # Cor de fundo pela criticidade
        if hs < 40 or rpn > 100:
            pdf.set_fill_color(254, 202, 202)
            txt_color = (127, 29, 29)
            bold = True
        elif hs < 60 or rpn > 50:
            pdf.set_fill_color(254, 240, 138)
            txt_color = (120, 53, 15)
            bold = True
        else:
            pdf.set_fill_color(255, 255, 255)
            txt_color = (30, 41, 59)
            bold = False

        pdf.set_text_color(*txt_color)
        if bold:
            pdf.set_font('Arial', 'B', 7)
        else:
            pdf.set_font('Arial', '', 7)

        # Ação sugerida resumida (primeira linha do diagnóstico, ~40 chars)
        diag = str(row.get('DIAGNOSTICO_IA', ''))
        primeira_linha = diag.split('\n')[0] if diag else '-'
        # Remove emojis do início (que não imprimem em latin-1)
        primeira_linha_limpa = ''.join(c for c in primeira_linha if ord(c) < 128)[:42].strip()
        if not primeira_linha_limpa:
            primeira_linha_limpa = 'Verificar laudo'

        valores = [
            str(i),
            _safe_latin(str(row.get('FROTA', '-')))[:12],
            _safe_latin(str(row.get('COMPARTIMENTO', '-')))[:14],
            _safe_latin(str(row.get('Família do equipamento', '-')))[:10],
            f'{hs:.0f}',
            f'{rpn:.0f}',
            _safe_latin(tend),
            _safe_latin(primeira_linha_limpa),
        ]
        for w, v in zip(larguras, valores):
            align = 'C' if w < 20 else 'L'
            pdf.cell(w, 5, _safe_latin(str(v)), 1, 0, align, fill=True)
        pdf.ln()

    # ==========================================
    # DISTRIBUIÇÃO DO HEALTH SCORE — visual autoexplicativo
    # ==========================================
    pdf.ln(4)
    pdf.set_font('Arial', 'B', 10)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 5, _safe_latin('Distribuição da Saúde da Frota'), 0, 1, 'L')

    pdf.set_font('Arial', '', 7)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(0, 4,
              _safe_latin('Cada amostra recebe uma nota de 0 a 100 baseada na química do óleo e na classificação do laboratório ALS. Quanto MAIOR a nota, MELHOR a condição do equipamento.'),
              0, 1, 'L')
    pdf.ln(2)

    if 'HEALTH_SCORE' in df_pdf.columns and len(df_pdf) > 0:
        faixas = [
            {
                'label': 'EXCELENTE',
                'range': '80 a 100',
                'descricao': 'Operando dentro dos parâmetros normais. Nenhuma ação necessária.',
                'n': int((df_pdf['HEALTH_SCORE'] >= 80).sum()),
                'cor_fundo': (220, 252, 231),
                'cor_borda': (22, 163, 74),
                'cor_texto': (22, 101, 52),
                'letra': 'A',
            },
            {
                'label': 'BOM',
                'range': '60 a 79',
                'descricao': 'Pequenos sinais de desgaste dentro do esperado. Manter monitoramento.',
                'n': int(((df_pdf['HEALTH_SCORE'] >= 60) & (df_pdf['HEALTH_SCORE'] < 80)).sum()),
                'cor_fundo': (254, 243, 199),
                'cor_borda': (134, 239, 172),
                'cor_texto': (22, 101, 52),
                'letra': 'B',
            },
            {
                'label': 'ATENÇÃO',
                'range': '40 a 59',
                'descricao': 'Anomalia confirmada pelo laboratório. Programar inspeção na próxima parada.',
                'n': int(((df_pdf['HEALTH_SCORE'] >= 40) & (df_pdf['HEALTH_SCORE'] < 60)).sum()),
                'cor_fundo': (254, 215, 170),
                'cor_borda': (245, 158, 11),
                'cor_texto': (124, 45, 18),
                'letra': 'C',
            },
            {
                'label': 'CRÍTICO',
                'range': '0 a 39',
                'descricao': 'Risco operacional alto. Intervenção imediata recomendada.',
                'n': int((df_pdf['HEALTH_SCORE'] < 40).sum()),
                'cor_fundo': (254, 202, 202),
                'cor_borda': (220, 38, 38),
                'cor_texto': (127, 29, 29),
                'letra': 'D',
            },
        ]
        total = len(df_pdf)

        # 1) Barra horizontal proporcional no topo (visual rápido)
        y_bar = pdf.get_y()
        x_bar = 10
        bar_w_total = 190
        x_curr = x_bar
        for fx in faixas:
            if fx['n'] > 0:
                w = (fx['n'] / total) * bar_w_total
                pdf.set_fill_color(*fx['cor_borda'])
                pdf.rect(x_curr, y_bar, w, 6, 'F')
                if w > 22:
                    pdf.set_xy(x_curr, y_bar)
                    pdf.set_text_color(255, 255, 255)
                    pdf.set_font('Arial', 'B', 7)
                    pdf.cell(w, 6,
                              _safe_latin(f"{fx['n']} ({100*fx['n']/total:.0f}%)"),
                              0, 0, 'C')
                x_curr += w
        pdf.set_y(y_bar + 9)

        # 2) Cards autoexplicativos (2 colunas × 2 linhas)
        col_w = 93
        card_h = 22
        gap_x = 4
        positions = [
            (10,            y_bar + 9),
            (10 + col_w + gap_x, y_bar + 9),
            (10,            y_bar + 9 + card_h + 2),
            (10 + col_w + gap_x, y_bar + 9 + card_h + 2),
        ]

        for fx, (x, y) in zip(faixas, positions):
            # Card com fundo claro e borda colorida
            pdf.set_fill_color(*fx['cor_fundo'])
            pdf.set_draw_color(*fx['cor_borda'])
            pdf.rect(x, y, col_w, card_h, 'DF')

            # Faixa colorida lateral esquerda (mais grossa)
            pdf.set_fill_color(*fx['cor_borda'])
            pdf.rect(x, y, 3, card_h, 'F')

            # Círculo com letra (selo)
            cx = x + 11
            cy = y + 10
            pdf.set_fill_color(*fx['cor_borda'])
            pdf.ellipse(cx - 4, cy - 4, 8, 8, 'F')
            pdf.set_xy(cx - 4, cy - 3.5)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(8, 6, fx['letra'], 0, 0, 'C')

            # Texto: rótulo + range + qtd
            pdf.set_xy(x + 19, y + 2)
            pdf.set_font('Arial', 'B', 9)
            pdf.set_text_color(*fx['cor_texto'])
            pdf.cell(col_w - 22, 4,
                      _safe_latin(f"{fx['label']}   (nota {fx['range']})"),
                      0, 1, 'L')

            # Quantidade (destaque)
            pdf.set_xy(x + 19, y + 7)
            pdf.set_font('Arial', 'B', 11)
            pdf.set_text_color(*fx['cor_borda'])
            pct = (100 * fx['n'] / total) if total > 0 else 0
            pdf.cell(col_w - 22, 5,
                      _safe_latin(f"{fx['n']} amostra{'s' if fx['n'] != 1 else ''}   |   {pct:.0f}% do total"),
                      0, 1, 'L')

            # Descrição
            pdf.set_xy(x + 19, y + 13)
            pdf.set_font('Arial', '', 6.5)
            pdf.set_text_color(*fx['cor_texto'])
            pdf.multi_cell(col_w - 22, 3, _safe_latin(fx['descricao']), 0, 'L')

        pdf.set_y(y_bar + 9 + 2 * (card_h + 2) + 2)

    # Nota: A página 2 (Análise Estratégica com heatmap/pareto/benchmark) foi
    # removida a pedido do usuário. As funções helpers _gerar_heatmap_*,
    # _gerar_pareto_* e _gerar_benchmark_* permanecem disponíveis para uso
    # futuro caso seja necessário reativar.


def _checklist_acoes_oficina(diagnostico_str):
    """
    Mapeia o diagnóstico IA em um checklist prático para a oficina.
    Retorna lista de tuplas (item, tempo_min, ferramenta).
    """
    checklist = []
    diag = str(diagnostico_str).lower()

    # Mapas de problema → ações concretas
    if 'silício' in diag or 'silicio' in diag or 'poeira' in diag:
        checklist.append(('Inspecionar visualmente filtro de ar', 5, 'Visual'))
        checklist.append(('Verificar assento da tampa do filtro', 5, 'Manual'))
        checklist.append(('Apertar abraçadeiras dos mangotes', 10, 'Chave de fenda'))
        checklist.append(('Inspecionar mangotes (rachaduras/folga)', 8, 'Visual'))
    if 'ferro' in diag and 'desgaste' in diag:
        checklist.append(('Abrir filtro de óleo para análise de partículas', 15, 'Chave de filtro'))
        checklist.append(('Inspecionar bujão magnético', 5, 'Chave fixa'))
        checklist.append(('Verificar nível de óleo no cárter', 3, 'Manual'))
    if 'cobre' in diag or 'bronzina' in diag:
        checklist.append(('Inspecionar bronzinas/mancais', 30, 'Kit oficina'))
        checklist.append(('Verificar pressão do óleo', 5, 'Manômetro'))
    if 'sódio' in diag or 'potássio' in diag or 'arrefecimento' in diag:
        checklist.append(('Pressurizar sistema de arrefecimento', 15, 'Compressor'))
        checklist.append(('Verificar trocador de calor (vazamentos)', 20, 'Visual + UV'))
        checklist.append(('Conferir nível do reservatório', 3, 'Manual'))
    if 'diluição' in diag or 'diesel' in diag or 'combustível' in diag:
        checklist.append(('Inspecionar bicos injetores', 25, 'Kit injeção'))
        checklist.append(('Testar bomba injetora', 30, 'Banca de teste'))
    if 'fuligem' in diag or 'queima' in diag:
        checklist.append(('Verificar filtro de ar (admissão restrita?)', 5, 'Visual'))
        checklist.append(('Inspecionar bicos injetores (entupimento)', 25, 'Kit injeção'))
        checklist.append(('Verificar EGR/turbo', 20, 'Kit oficina'))
    if 'tbn' in diag or 'neutraliz' in diag:
        checklist.append(('Programar troca de óleo IMEDIATA', 60, 'Kit troca'))
    if 'pq' in diag or 'fadiga' in diag:
        checklist.append(('Análise de detritos do filtro (laboratório)', 0, 'Laboratório'))
        checklist.append(('Considerar desmontagem para inspeção', 240, 'Oficina'))

    return checklist[:6]  # máximo 6 itens para caber no card


def _renderizar_checklist_oficina(pdf, diagnostico_str, x_start, y_start, max_w=190):
    """Renderiza o checklist como uma tabela compacta no PDF."""
    itens = _checklist_acoes_oficina(diagnostico_str)
    if not itens:
        return y_start

    pdf.set_xy(x_start, y_start)
    pdf.set_font('Arial', 'B', 8)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(max_w, 5, _safe_latin('🔧 CHECKLIST SUGERIDO PARA A OFICINA'), 0, 1, 'L')

    # Cabeçalho
    pdf.set_fill_color(241, 245, 249)
    pdf.set_text_color(71, 85, 105)
    pdf.set_font('Arial', 'B', 6.5)
    pdf.set_x(x_start)
    pdf.cell(8, 4, _safe_latin('OK'), 1, 0, 'C', fill=True)
    pdf.cell(115, 4, _safe_latin('Item'), 1, 0, 'L', fill=True)
    pdf.cell(20, 4, _safe_latin('Tempo'), 1, 0, 'C', fill=True)
    pdf.cell(47, 4, _safe_latin('Ferramenta'), 1, 1, 'C', fill=True)

    pdf.set_font('Arial', '', 7)
    pdf.set_text_color(30, 41, 59)
    for item, tempo_min, ferr in itens:
        pdf.set_x(x_start)
        pdf.cell(8, 4.5, '', 1, 0, 'C')  # checkbox vazio
        pdf.cell(115, 4.5, _safe_latin(item)[:65], 1, 0, 'L')
        tempo_str = f'{tempo_min} min' if tempo_min > 0 else '-'
        pdf.cell(20, 4.5, tempo_str, 1, 0, 'C')
        pdf.cell(47, 4.5, _safe_latin(ferr)[:24], 1, 1, 'C')

    return pdf.get_y()


def _renderizar_tabela_quimica_completa(pdf, row, x_start=10, y_start=None):
    """
    Renderiza um bloco completo de química na vertical, agrupado em 5 categorias.
    Cada célula é colorida pelo status do elemento (verde/amarelo/vermelho).
    Retorna o y final.
    """
    if y_start is None:
        y_start = pdf.get_y()

    # Estrutura: (titulo_grupo, [(label, key, unidade, casas_dec), ...])
    grupos = [
        ('METAIS DE DESGASTE (ppm)', [
            ('Fe', 'Ferro', 0), ('Cr', 'Cromo', 0), ('Ni', 'Níquel', 0),
            ('Pb', 'Chumbo', 0), ('Sn', 'Estanho', 0), ('Cu', 'Cobre', 0),
            ('Al', 'Alumínio', 0), ('Mn', 'Manganês', 0),
        ]),
        ('CONTAMINANTES (ppm)', [
            ('Si', 'Silício', 0), ('Na', 'Sódio', 0),
            ('K',  'Potássio', 0), ('B', 'Boro', 0),
        ]),
        ('ADITIVOS DO ÓLEO (ppm)', [
            ('Mg', 'Magnésio', 0), ('Ca', 'Cálcio', 0), ('P', 'Fósforo', 0),
            ('Zn', 'Zinco', 0), ('Ba', 'Bário', 0), ('Mo', 'Molibdênio', 0),
        ]),
        ('CONDIÇÃO FÍSICO-QUÍMICA', [
            ('Visc. 100C', 'Viscosidade', 2),
            ('TBN', 'TBN', 2),
            ('Diluiç. Diesel %', 'Diluição Diesel', 1),
            ('Água (ppm)', 'KF_Agua', 0),
            ('PQ Index', 'INDICE_PQ', 0),
        ]),
        ('ANÁLISES FTIR', [
            ('Fuligem', 'Fuligem', 2),
            ('Oxidação', 'Oxidação_FTIR', 1),
            ('Nitração', 'Nitração_FTIR', 1),
            ('Sulfatação', 'Sulfatação_FTIR', 1),
            ('Glicol', 'Glicol_FTIR', 1),
        ]),
    ]

    pdf.set_xy(x_start, y_start)
    largura_total = 190
    altura_titulo = 4.5
    altura_linha = 4.5

    for titulo, elementos in grupos:
        # Título do grupo
        pdf.set_fill_color(30, 41, 59)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Arial', 'B', 7)
        pdf.set_x(x_start)
        pdf.cell(largura_total, altura_titulo, _safe_latin('  ' + titulo), 0, 1, 'L', fill=True)

        # Linha com os elementos (até 8 por linha)
        elementos_validos = [(lbl, key, dec) for lbl, key, dec in elementos]
        n = len(elementos_validos)
        if n == 0:
            continue

        cell_w = largura_total / n

        # Linha de labels
        pdf.set_fill_color(241, 245, 249)
        pdf.set_text_color(71, 85, 105)
        pdf.set_font('Arial', 'B', 6)
        pdf.set_x(x_start)
        for lbl, _, _ in elementos_validos:
            pdf.cell(cell_w, altura_linha, _safe_latin(lbl), 1, 0, 'C', fill=True)
        pdf.ln()

        # Linha de valores (com cor de status por célula)
        pdf.set_font('Arial', '', 7)
        pdf.set_x(x_start)
        for _, key, dec in elementos_validos:
            valor = row.get(key, 0)
            try:
                valor_num = float(valor) if pd.notna(valor) else 0.0
            except (ValueError, TypeError):
                valor_num = 0.0

            # Determina cor do fundo
            r, g, b, status = _cor_status_elemento(key, valor_num)
            pdf.set_fill_color(r, g, b)

            # Cor do texto: mais escuro quando crítico
            if status == 'critico':
                pdf.set_text_color(127, 29, 29)
                pdf.set_font('Arial', 'B', 7)
            elif status == 'alerta':
                pdf.set_text_color(120, 53, 15)
                pdf.set_font('Arial', 'B', 7)
            else:
                pdf.set_text_color(30, 41, 59)
                pdf.set_font('Arial', '', 7)

            # Formato do valor
            if valor_num == 0 and status == 'sem_ref':
                txt = '-'
            elif dec == 0:
                txt = f'{valor_num:.0f}'
            else:
                txt = f'{valor_num:.{dec}f}'

            pdf.cell(cell_w, altura_linha, txt, 1, 0, 'C', fill=True)
        pdf.ln()

    # Restaura cores padrão
    pdf.set_text_color(0, 0, 0)
    pdf.set_fill_color(255, 255, 255)
    return pdf.get_y()


class PDFPlanoAcaoALS(FPDF):
    def header(self):
        caminho_logo = os.path.join(os.path.dirname(os.path.dirname(__file__)), "logo_cedro.png")
        if os.path.exists(caminho_logo):
            self.image(caminho_logo, 10, 8, 15)
        self.set_font('Arial', 'B', 16)
        self.set_text_color(30, 41, 59)
        self.set_xy(30, 10)
        self.cell(0, 8, 'RELATORIO DE ANALISE PREDITIVA E ORDEM DE SERVICO', 0, 1, 'L')
        self.set_font('Arial', '', 9)
        self.set_text_color(100, 116, 139)
        self.set_x(30)
        self.cell(0, 4, 'Gestao de Frotas | Extracao do Laboratorio', 0, 1, 'L')
        self.set_draw_color(226, 232, 240)
        self.line(10, 24, 200, 24)
        self.set_y(28)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(150, 150, 150)
        self.cell(0, 10, f'Gerado via Sistema Cedro | Pagina {self.page_no()}', 0, 0, 'C')


@st.cache_data(show_spinner="A gerar PDF de Ações (Gerencial + OS)...", ttl=60)
def gerar_pdf_plano_acao(df_pdf, df_full):
    arquivos_temp = []
    try:
        pdf = PDFPlanoAcaoALS(orientation='P', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=15)

        # ==========================================
        # PÁGINAS 1-2: NOVA CAPA EXECUTIVA
        # KPIs + Top 10 prioritárias + Heatmap + Pareto + Benchmark fabricantes
        # ==========================================
        _gerar_capa_executiva(pdf, df_pdf, arquivos_temp)

        pdf.add_page()

        # ==========================================
        # PÁGINAS DE ORDENS DE SERVIÇO (CARDS) E TIMELINE DUPLA
        # ==========================================
        for _, row in df_pdf.iterrows():
            start_y = pdf.get_y()
            # Margem de segurança maior devido ao gráfico duplo
            if start_y > 100:
                pdf.add_page()
                start_y = pdf.get_y()

            # --- HEADER DA O.S. ---
            status_v = row.get('STATUS_CORRIGIDO', '')
            if status_v == 'Crítico':
                bg_r, bg_g, bg_b = 239, 68, 68
                txt_status = "CRITICO  STOP"
            elif status_v == 'Alerta':
                bg_r, bg_g, bg_b = 245, 158, 11
                txt_status = "ATENCAO  !"
            else:
                bg_r, bg_g, bg_b = 16, 185, 129
                txt_status = "NORMAL  OK"

            pdf.set_fill_color(bg_r, bg_g, bg_b)
            pdf.rect(140, start_y, 60, 10, 'F')
            pdf.set_xy(140, start_y + 2)
            pdf.set_font('Arial', 'B', 12)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(60, 6, txt_status, 0, 0, 'C')

            # --- INFOGRÁFICO: RÉGUA DE HISTÓRICO ---
            frota_alvo = row.get('FROTA')
            comp_alvo = row.get('COMPARTIMENTO')
            df_hist = df_full[(df_full['FROTA'] == frota_alvo) & (df_full['COMPARTIMENTO'] == comp_alvo)].copy()
            df_hist = df_hist.dropna(subset=['DATA_COLETA']).sort_values(by='DATA_COLETA').tail(5)

            if not df_hist.empty and len(df_hist) > 0:
                pdf.set_xy(140, start_y + 11)
                pdf.set_font('Arial', 'B', 6)
                pdf.set_text_color(100, 116, 139)
                pdf.cell(60, 3, "HISTORICO RECENTE:", 0, 1, 'R')

                q_width = 8
                q_space = 2
                total_w = len(df_hist) * (q_width + q_space) - q_space
                start_q_x = 200 - total_w

                y_q = start_y + 15
                for _, h_row in df_hist.iterrows():
                    h_stat = h_row.get('STATUS_CORRIGIDO', '')
                    if h_stat == 'Crítico':
                        pdf.set_fill_color(239, 68, 68)
                    elif h_stat == 'Alerta':
                        pdf.set_fill_color(245, 158, 11)
                    else:
                        pdf.set_fill_color(16, 185, 129)

                    pdf.rect(start_q_x, y_q, q_width, 4, 'F')
                    start_q_x += (q_width + q_space)

            # --- DADOS DO EQUIPAMENTO E CLIENTE ---
            pdf.set_font('Arial', 'B', 8)
            pdf.set_text_color(71, 85, 105)
            pdf.set_xy(10, start_y)

            cliente_str = _safe_latin(str(row.get('CLIENTE', '')))[:45]
            obra_str = _safe_latin(str(row.get('OBRA', '')))[:45]
            amostra_str = _safe_latin(str(row.get('NUM_AMOSTRA', '')))
            dt_str = row['DATA_COLETA'].strftime('%d/%m/%Y') if pd.notnull(row.get('DATA_COLETA')) else "-"

            pdf.cell(100, 4, f"CLIENTE: {cliente_str}", 0, 1, 'L')
            pdf.cell(100, 4, f"UNIDADE/OBRA: {obra_str}", 0, 1, 'L')
            pdf.cell(100, 4, f"AMOSTRA: {amostra_str}  |  DATA DA COLETA: {dt_str}", 0, 1, 'L')

            # Linha extra com responsável ALS e plano
            resp_als_str = _safe_latin(str(row.get('RESPONSAVEL_ALS', '-')))[:45]
            plano_str = _safe_latin(str(row.get('PLANO_ANALISE', '-')))[:30]
            pdf.set_font('Arial', 'I', 7)
            pdf.set_text_color(100, 116, 139)
            pdf.cell(100, 3.5, f"Analista ALS: {resp_als_str}  |  Plano: {plano_str}", 0, 1, 'L')

            box_equip_y = start_y + 17
            pdf.set_fill_color(248, 250, 252)
            pdf.set_draw_color(226, 232, 240)
            pdf.rect(10, box_equip_y, 125, 22, 'DF')

            pdf.set_xy(12, box_equip_y + 2)
            pdf.set_font('Arial', 'B', 9)
            pdf.set_text_color(15, 23, 42)

            frota_str = _safe_latin(str(row.get('FROTA', '')))
            comp_str = _safe_latin(str(row.get('COMPARTIMENTO', '')))
            fam_str = _safe_latin(str(row.get('Família do equipamento', '')))
            fabr_str = _safe_latin(str(row.get('FABRICANTE', '-')))
            mod_str = _safe_latin(str(row.get('MODELO', '')))
            visc_oleo = _safe_latin(str(row.get('VISCOSIDADE_OLEO', '-')))
            hr_oleo = str(row.get('HORAS_OLEO', '0'))
            hr_eq = str(row.get('HORAS_EQUIP', '0'))
            tr_oleo = str(row.get('TROCOU_OLEO_FLAG', 'Não'))

            pdf.cell(60, 5, f"TAG/FROTA: {frota_str}", 0, 0, 'L')
            pdf.cell(60, 5, f"COMPARTIMENTO: {comp_str}", 0, 1, 'L')

            pdf.set_x(12)
            pdf.set_font('Arial', '', 8)
            pdf.set_text_color(71, 85, 105)
            pdf.cell(120, 4, f"FABRIC./FAMILIA/MODELO: {fabr_str} - {fam_str} - {mod_str}", 0, 1, 'L')

            pdf.set_x(12)
            pdf.cell(60, 4, f"HR EQUIP: {hr_eq} h  |  HR OLEO: {hr_oleo} h", 0, 0, 'L')
            pdf.cell(60, 4, f"OLEO: {visc_oleo}  |  TROCA: {tr_oleo}", 0, 1, 'L')

            av_y = box_equip_y + 24
            pdf.set_xy(10, av_y)
            pdf.set_font('Arial', 'B', 8)
            pdf.set_text_color(30, 41, 59)
            pdf.cell(190, 5, "AVALIACAO:", 0, 1, 'L')

            pdf.set_x(10)
            pdf.set_font('Arial', '', 8)
            av_str = _safe_latin(str(row.get('AVALIACAO', '')))
            pdf.multi_cell(190, 4, av_str, 0, 'L')

            ac_y = pdf.get_y() + 2
            pdf.set_xy(10, ac_y)
            pdf.set_font('Arial', 'B', 8)
            pdf.set_text_color(30, 41, 59)
            pdf.cell(190, 5, "ACOES DE INSPECAO:", 0, 1, 'L')

            pdf.set_x(10)
            pdf.set_font('Arial', '', 8)
            ac_str = str(row.get('ACOES_INSPECAO', '')).encode('latin-1', 'ignore').decode('latin-1')
            pdf.multi_cell(190, 4, ac_str, 0, 'L')

            qui_y = pdf.get_y() + 3

            # ==========================================
            # NOVO: Bloco completo de química (5 grupos com 22+ elementos)
            # Substitui a antiga tabela de 5 células
            # ==========================================
            qui_end_y = _renderizar_tabela_quimica_completa(pdf, row, x_start=10, y_start=qui_y)
            pdf.set_y(qui_end_y + 2)

            # ==========================================
            # NOVO: Radar de desgaste para amostras críticas/alerta
            # ==========================================
            status_row = row.get('STATUS_CORRIGIDO', 'Normal')
            if MATPLOTLIB_AVAILABLE and status_row in ('Crítico', 'Alerta'):
                try:
                    tmp_radar = tempfile.NamedTemporaryFile(
                        prefix="cedro_oleo_radar_", suffix=".png", delete=False
                    )
                    radar_path = tmp_radar.name
                    tmp_radar.close()
                    arquivos_temp.append(radar_path)

                    if _gerar_radar_desgaste(row, radar_path):
                        # Verifica espaço; se faltar, nova página
                        if pdf.get_y() > 200:
                            pdf.add_page()
                        # Radar ocupa 70mm de largura, alinhado à direita; texto de elementos ALS à esquerda
                        radar_y = pdf.get_y() + 2
                        pdf.image(radar_path, x=130, y=radar_y, w=70)

                        # Bloco de "Elementos em alerta segundo ALS" ao lado esquerdo do radar
                        pdf.set_xy(10, radar_y + 4)
                        pdf.set_font('Arial', 'B', 8)
                        pdf.set_text_color(127, 29, 29)
                        pdf.cell(115, 5, _safe_latin('🚩 ELEMENTOS SINALIZADOS PELA ALS'), 0, 1, 'L')

                        elem_als = str(row.get('ELEMENTOS_EM_ALERTA_ALS', '')).strip()
                        if elem_als and elem_als != 'nan':
                            pdf.set_xy(10, pdf.get_y() + 1)
                            pdf.set_font('Arial', '', 8)
                            pdf.set_text_color(30, 41, 59)
                            pdf.multi_cell(115, 4, _safe_latin('• ' + elem_als), 0, 'L')
                        else:
                            pdf.set_xy(10, pdf.get_y() + 1)
                            pdf.set_font('Arial', 'I', 7)
                            pdf.set_text_color(100, 116, 139)
                            pdf.multi_cell(115, 4, _safe_latin('Nenhum elemento individualmente sinalizado pela ALS — verificar conjunto.'), 0, 'L')

                        # Diagnóstico sugerido (sintético)
                        pdf.set_xy(10, pdf.get_y() + 2)
                        pdf.set_font('Arial', 'B', 8)
                        pdf.set_text_color(30, 41, 59)
                        pdf.cell(115, 5, _safe_latin('DIAGNÓSTICO SUGERIDO'), 0, 1, 'L')
                        diag_str = _safe_latin(str(row.get('DIAGNOSTICO_IA', '-')))
                        pdf.set_xy(10, pdf.get_y() + 1)
                        pdf.set_font('Arial', '', 7)
                        pdf.set_text_color(71, 85, 105)
                        pdf.multi_cell(115, 3.5, diag_str, 0, 'L')

                        # Avança Y para depois do radar (que tem ~70mm de altura)
                        pdf.set_y(max(pdf.get_y(), radar_y + 72))
                except Exception as e:
                    logger.warning("Radar/diagnóstico não gerado: %s", e)
                    try:
                        plt.close('all')
                    except Exception:
                        pass

            # Nota: O checklist da oficina foi removido a pedido do usuário.
            # A função _renderizar_checklist_oficina permanece disponível para
            # uso futuro caso seja necessário reativar.

            box_acao_y = pdf.get_y() + 4
            pdf.set_fill_color(255, 255, 255)
            pdf.set_draw_color(100, 116, 139)
            pdf.rect(10, box_acao_y, 190, 20, 'D')

            pdf.set_xy(12, box_acao_y + 1)
            pdf.set_font('Arial', 'B', 7)
            pdf.set_text_color(100, 116, 139)
            pdf.cell(190, 4, "PLANO DE ACAO E RETORNO DA OFICINA (A Preencher):", 0, 1, 'L')

            acao_bd = str(row.get('ACAO_GESTAO', '')).strip()
            if acao_bd and acao_bd.upper() != 'NAN':
                pdf.set_xy(12, box_acao_y + 6)
                pdf.set_font('Arial', 'I', 8)
                pdf.set_text_color(30, 41, 59)
                pdf.multi_cell(186, 4, _safe_latin(acao_bd), 0, 'L')
            else:
                pdf.set_draw_color(226, 232, 240)
                pdf.line(12, box_acao_y + 9, 198, box_acao_y + 9)
                pdf.line(12, box_acao_y + 14, 198, box_acao_y + 14)
                pdf.line(12, box_acao_y + 19, 198, box_acao_y + 19)

            pdf.set_y(box_acao_y + 24)

            # --- INSERÇÃO DE GRÁFICO DUPLO NO PDF (PPM + STATUS TIMELINE COM P&B SEGURO) ---
            if MATPLOTLIB_AVAILABLE:
                try:
                    if not df_hist.empty and len(df_hist) > 0:
                        elementos_analise = ['Ferro', 'Silício', 'Cobre', 'Sódio', 'Potássio', 'Fuligem', 'INDICE_PQ']
                        plot_cols = [c for c in elementos_analise if c in df_hist.columns and df_hist[c].sum() > 0]

                        if plot_cols:
                            # Criação do painel duplo (Dois gráficos em um, compartilhando o eixo X/Datas)
                            fig_mini, (ax_ppm, ax_status) = plt.subplots(
                                nrows=2, ncols=1, figsize=(8, 3.5), sharex=True,
                                gridspec_kw={'height_ratios': [2, 1.2]}
                            )

                            # --- GRÁFICO SUPERIOR: Evolução Química (PPM) ---
                            line_styles = ['-', '--', '-.', ':', (0, (3, 1, 1, 1))]  # Diferentes traços para P&B
                            last_points = []

                            for idx, col in enumerate(plot_cols[:5]):
                                valid_data = df_hist[['DATA_COLETA', col]].dropna()
                                if not valid_data.empty:
                                    style = line_styles[idx % len(line_styles)]
                                    line_obj = ax_ppm.plot(df_hist['DATA_COLETA'], df_hist[col],
                                                           marker='o', label=col, markersize=3,
                                                           linewidth=1.5, linestyle=style)[0]

                                    last_x_date = valid_data['DATA_COLETA'].iloc[-1]
                                    last_x_num = mdates.date2num(last_x_date)
                                    last_y = valid_data[col].iloc[-1]

                                    last_points.append({
                                        'col': col, 'x_date': last_x_date, 'x_num': last_x_num,
                                        'y': last_y, 'color': line_obj.get_color()
                                    })

                            # Anti-Sobreposição dos Rótulos (para P&B ser legível)
                            if last_points:
                                last_points.sort(key=lambda p: p['y'])
                                fig_mini.canvas.draw()

                                y_pixels = []
                                for pt in last_points:
                                    coord = ax_ppm.transData.transform((pt['x_num'], pt['y']))
                                    y_pixels.append(coord[1])

                                min_pixel_dist = 14.0
                                for _ in range(20):
                                    for i in range(len(y_pixels) - 1):
                                        if (y_pixels[i + 1] - y_pixels[i]) < min_pixel_dist:
                                            overlap = min_pixel_dist - (y_pixels[i + 1] - y_pixels[i])
                                            y_pixels[i] -= overlap / 2.0
                                            y_pixels[i + 1] += overlap / 2.0

                                for i, pt in enumerate(last_points):
                                    orig_y_pixel = ax_ppm.transData.transform((pt['x_num'], pt['y']))[1]
                                    offset_y = y_pixels[i] - orig_y_pixel

                                    ax_ppm.annotate(f" {pt['col']}",
                                                    xy=(pt['x_date'], pt['y']), xytext=(8, offset_y),
                                                    textcoords="offset points", fontsize=7, fontweight='bold',
                                                    color=pt['color'], va='center',
                                                    arrowprops=dict(arrowstyle="-", color=pt['color'], lw=0.5,
                                                                    alpha=0.6) if abs(offset_y) > 3 else None)

                            left, right = ax_ppm.get_xlim()
                            ax_ppm.set_xlim(left, right + (right - left) * 0.20)
                            ax_ppm.set_title(f"Evolucao Quimica (PPM): {comp_alvo} - {frota_alvo}", fontsize=8,
                                             color='#333333', loc='left', pad=3)
                            ax_ppm.tick_params(axis='both', labelsize=6)
                            ax_ppm.grid(True, linestyle='--', alpha=0.3)
                            ax_ppm.legend(fontsize=6, loc='center left', bbox_to_anchor=(1, 0.5))
                            ax_ppm.spines['top'].set_visible(False);
                            ax_ppm.spines['right'].set_visible(False)

                            # --- GRÁFICO INFERIOR: Timeline de Status e Troca de Óleo ---
                            df_hist['STATUS_NUM'] = df_hist['STATUS_CORRIGIDO'].map(
                                {'Normal': 1, 'Alerta': 2, 'Crítico': 3})

                            # Linha conectora de fundo
                            ax_status.plot(df_hist['DATA_COLETA'], df_hist['STATUS_NUM'], color='gray', linestyle='-',
                                           linewidth=1, alpha=0.4)

                            # Plot de Status com formas distintas para P&B (Círculo, Triângulo, Quadrado)
                            for s_nome, s_val, s_cor, s_marker in [('Normal', 1, '#10B981', 'o'),
                                                                   ('Alerta', 2, '#F59E0B', '^'),
                                                                   ('Crítico', 3, '#EF4444', 's')]:
                                mask = df_hist['STATUS_CORRIGIDO'] == s_nome
                                if mask.any():
                                    ax_status.scatter(df_hist.loc[mask, 'DATA_COLETA'], df_hist.loc[mask, 'STATUS_NUM'],
                                                      color=s_cor, marker=s_marker, s=40, label=s_nome, zorder=5)

                            # Destaque para a Troca de Óleo
                            mask_oleo = df_hist['TROCOU_OLEO_FLAG'] == 'Sim'
                            if mask_oleo.any():
                                ax_status.scatter(df_hist.loc[mask_oleo, 'DATA_COLETA'],
                                                  df_hist.loc[mask_oleo, 'STATUS_NUM'],
                                                  facecolors='none', edgecolors='black', marker='D', s=80,
                                                  linewidths=1.5, label='Troca de Oleo', zorder=6)
                                for _, row_oil in df_hist[mask_oleo].iterrows():
                                    ax_status.annotate("Troca", xy=(row_oil['DATA_COLETA'], row_oil['STATUS_NUM']),
                                                       xytext=(0, 12), textcoords="offset points", ha='center',
                                                       fontsize=6, fontweight='bold', color='black')

                            ax_status.set_yticks([1, 2, 3])
                            ax_status.set_yticklabels(['Normal', 'Alerta', 'Critico'], fontsize=7, fontweight='bold')
                            ax_status.set_ylim(0.5, 3.8)  # Um pouco mais de espaço para o texto "Troca"
                            ax_status.grid(True, axis='y', linestyle='--', alpha=0.3)
                            ax_status.legend(fontsize=6, loc='center left', bbox_to_anchor=(1, 0.5))
                            ax_status.spines['top'].set_visible(False);
                            ax_status.spines['right'].set_visible(False)
                            ax_status.set_title("Timeline de Saude (Status)", fontsize=8, color='#333333', loc='left',
                                                pad=3)

                            # Formatação da Data (Aplica no eixo inferior automaticamente pelo sharex)
                            ax_status.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m/%y'))
                            fig_mini.autofmt_xdate(rotation=0, ha='center')

                            plt.tight_layout()

                            tmp_trend = tempfile.NamedTemporaryFile(prefix="cedro_oleo_", suffix=".png", delete=False)
                            img_trend = tmp_trend.name
                            tmp_trend.close()
                            arquivos_temp.append(img_trend)

                            fig_mini.savefig(img_trend, dpi=150, bbox_inches='tight')
                            plt.close(fig_mini)

                            if pdf.get_y() > 200:  # Ajustado pois a imagem agora é mais alta
                                pdf.add_page()

                            pdf.image(img_trend, x=10, y=pdf.get_y() + 2, w=190)
                            pdf.set_y(pdf.get_y() + 85)  # Deslocamento maior devido aos 2 gráficos

                except Exception as e:
                    # Bug 4 corrigido: loga o erro E deixa nota visível no PDF
                    logger.warning(
                        "Gráfico não gerado para %s / %s: %s",
                        frota_alvo, comp_alvo, e
                    )
                    pdf.set_font('Arial', 'I', 7)
                    pdf.set_text_color(150, 150, 150)
                    pdf.cell(
                        190, 5,
                        f"[Grafico nao disponivel para {str(frota_alvo).encode('latin-1','ignore').decode('latin-1')} / {str(comp_alvo).encode('latin-1','ignore').decode('latin-1')}]",
                        0, 1, 'C'
                    )
                    pdf.set_text_color(0, 0, 0)
                finally:
                    plt.close('all')

            pdf.set_draw_color(30, 41, 59)
            pdf.set_line_width(0.5)
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
            pdf.set_y(pdf.get_y() + 4)

        tmp_pdf = tempfile.NamedTemporaryFile(prefix="cedro_oleo_", suffix=".pdf", delete=False)
        pdf_path = tmp_pdf.name
        tmp_pdf.close()
        arquivos_temp.append(pdf_path)

        pdf.output(pdf_path)
        with open(pdf_path, "rb") as f:
            bytes_pdf = f.read()

        return bytes_pdf

    finally:
        if MATPLOTLIB_AVAILABLE:
            plt.close('all')

        for tmp_f in arquivos_temp:
            try:
                if os.path.exists(tmp_f):
                    os.remove(tmp_f)
            except:
                pass


# ==============================================================================
# ==============================================================================
# PERSISTÊNCIA DO DATASET
# Salva o dataset processado em disco para evitar re-upload a cada sessão.
# Cada novo upload faz MESCLA por NUM_AMOSTRA (amostras novas adicionadas,
# amostras com mesmo número são atualizadas pela versão mais recente).
#
# Estratégia: tenta usar parquet (mais rápido e leve), com fallback para pickle
# se pyarrow não estiver disponível no ambiente.
# ==============================================================================

_DIR_DADOS = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data')
CAMINHO_DATASET_PARQUET = os.path.join(_DIR_DADOS, 'oleo_amostras_atuais.parquet')
CAMINHO_DATASET_PICKLE  = os.path.join(_DIR_DADOS, 'oleo_amostras_atuais.pkl')
CAMINHO_DATASET_META    = os.path.join(_DIR_DADOS, 'oleo_amostras_meta.json')


def _testar_parquet_disponivel():
    """Verifica se pyarrow/fastparquet estão instalados."""
    try:
        import pyarrow  # noqa: F401
        return True
    except ImportError:
        try:
            import fastparquet  # noqa: F401
            return True
        except ImportError:
            return False


_PARQUET_OK = _testar_parquet_disponivel()


def carregar_dataset_persistido():
    """Carrega o dataset salvo em disco, se existir. Retorna (df, meta) ou (None, None)."""
    import json
    try:
        # Tenta parquet primeiro, depois pickle
        if _PARQUET_OK and os.path.exists(CAMINHO_DATASET_PARQUET):
            df = pd.read_parquet(CAMINHO_DATASET_PARQUET)
        elif os.path.exists(CAMINHO_DATASET_PICKLE):
            df = pd.read_pickle(CAMINHO_DATASET_PICKLE)
        else:
            return None, None
        meta = {}
        if os.path.exists(CAMINHO_DATASET_META):
            with open(CAMINHO_DATASET_META, 'r', encoding='utf-8') as f:
                meta = json.load(f)
        return df, meta
    except Exception as e:
        logger.warning("Falha ao carregar dataset persistido: %s", e)
        return None, None


def salvar_dataset_persistido(df, origem='upload'):
    """Salva o dataset no disco. Usa parquet se disponível, senão pickle."""
    import json
    try:
        os.makedirs(_DIR_DADOS, exist_ok=True)
        formato = 'pickle'
        try:
            if _PARQUET_OK:
                df.to_parquet(CAMINHO_DATASET_PARQUET, index=False)
                formato = 'parquet'
            else:
                df.to_pickle(CAMINHO_DATASET_PICKLE)
        except Exception:
            # Se parquet falhar por qualquer motivo, cai pro pickle
            df.to_pickle(CAMINHO_DATASET_PICKLE)
            formato = 'pickle'

        meta = {
            'atualizado_em':    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_amostras':   len(df),
            'frotas_distintas': int(df['FROTA'].nunique()) if 'FROTA' in df.columns else 0,
            'origem':           origem,
            'formato':          formato,
        }
        if 'DATA_COLETA' in df.columns and df['DATA_COLETA'].notna().any():
            meta['data_coleta_min'] = df['DATA_COLETA'].min().strftime('%d/%m/%Y')
            meta['data_coleta_max'] = df['DATA_COLETA'].max().strftime('%d/%m/%Y')
        with open(CAMINHO_DATASET_META, 'w', encoding='utf-8') as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        logger.exception("Falha ao salvar dataset persistido: %s", e)
        return False


def mesclar_datasets(df_novo, df_atual):
    """
    Mescla um novo lote (df_novo) com o dataset existente (df_atual).
    - Chave de identificação: NUM_AMOSTRA
    - Amostras com mesmo número são SUBSTITUÍDAS pela versão nova
    - Amostras antigas que não vêm no novo upload permanecem
    """
    if df_atual is None or df_atual.empty:
        return df_novo, {'novas': len(df_novo), 'atualizadas': 0, 'mantidas': 0}
    if df_novo is None or df_novo.empty:
        return df_atual, {'novas': 0, 'atualizadas': 0, 'mantidas': len(df_atual)}

    amostras_novas = set(df_novo['NUM_AMOSTRA'].astype(str))
    mask_substituidas = df_atual['NUM_AMOSTRA'].astype(str).isin(amostras_novas)
    n_atualizadas = mask_substituidas.sum()
    n_mantidas = (~mask_substituidas).sum()
    n_novas = len(df_novo) - n_atualizadas

    df_atual_filtrado = df_atual[~mask_substituidas].copy()
    df_resultado = pd.concat([df_atual_filtrado, df_novo], ignore_index=True, sort=False)

    return df_resultado, {
        'novas': int(n_novas),
        'atualizadas': int(n_atualizadas),
        'mantidas': int(n_mantidas),
    }


def limpar_dataset_persistido():
    """Remove o dataset salvo em disco (todos os formatos)."""
    try:
        for caminho in [CAMINHO_DATASET_PARQUET, CAMINHO_DATASET_PICKLE, CAMINHO_DATASET_META]:
            if os.path.exists(caminho):
                os.remove(caminho)
        return True
    except Exception as e:
        logger.warning("Falha ao limpar dataset: %s", e)
        return False


# ==============================================================================
# INTERFACE PRINCIPAL E FILTROS
# ==============================================================================

# Carrega dataset persistido na primeira visita à página
if 'dataset_oleo' not in st.session_state:
    df_persistido, meta_persistido = carregar_dataset_persistido()
    st.session_state['dataset_oleo'] = df_persistido
    st.session_state['dataset_meta'] = meta_persistido or {}

# Banner do estado atual do dataset
meta = st.session_state.get('dataset_meta', {})
if st.session_state['dataset_oleo'] is not None and meta:
    info_cols = st.columns([4, 1])
    with info_cols[0]:
        st.info(
            f"📦 **Dataset ativo** — {meta.get('total_amostras', 0)} amostras  |  "
            f"{meta.get('frotas_distintas', 0)} frotas  |  "
            f"Período: {meta.get('data_coleta_min', '-')} a {meta.get('data_coleta_max', '-')}  |  "
            f"Última atualização: {meta.get('atualizado_em', '-')}",
            icon="📦"
        )
    with info_cols[1]:
        if st.button("🗑️ Limpar dados", help="Remove o dataset persistido. Você precisará fazer novo upload."):
            if limpar_dataset_persistido():
                st.session_state['dataset_oleo'] = None
                st.session_state['dataset_meta'] = {}
                st.success("Dataset removido.")
                time.sleep(1)
                st.rerun()

# Bloco de upload — só expande quando não há dados persistidos
ja_tem_dados = st.session_state['dataset_oleo'] is not None
titulo_upload = "📂 Atualizar Laudo Laboratorial (Excel/CSV)" if ja_tem_dados else "📂 Importar Laudo Laboratorial (Excel/CSV)"

with st.expander(titulo_upload, expanded=not ja_tem_dados):
    if ja_tem_dados:
        st.caption(
            "💡 Seus dados já estão salvos. Use esta opção apenas para **adicionar novas amostras** "
            "ou **atualizar** amostras existentes. Amostras com mesmo número (NUM_AMOSTRA) "
            "serão substituídas pela versão nova; amostras antigas permanecem."
        )
    file_up = st.file_uploader("Faça upload do ficheiro exportado do laboratório", type=['xlsx', 'csv'])
    if file_up and st.button("Processar Laudos Químicos 🧪", type="primary"):
        file_bytes, file_name = _ler_bytes_upload(file_up)
        df_oleo = processar_laudo_oleo(file_bytes, file_name)
        if df_oleo is not None:
            df_atual = st.session_state['dataset_oleo']
            df_mesclado, stats = mesclar_datasets(df_oleo, df_atual)
            st.session_state['dataset_oleo'] = df_mesclado
            salvar_dataset_persistido(df_mesclado, origem=file_name)
            df_persistido, meta_persistido = carregar_dataset_persistido()
            st.session_state['dataset_meta'] = meta_persistido or {}

            # Feedback detalhado da mescla
            if stats['novas'] > 0 and stats['atualizadas'] > 0:
                st.success(
                    f"✅ Dataset atualizado! **{stats['novas']} novas** amostras, "
                    f"**{stats['atualizadas']} atualizadas**, "
                    f"**{stats['mantidas']} mantidas** do histórico."
                )
            elif stats['novas'] > 0:
                st.success(f"✅ {stats['novas']} novas amostras adicionadas ao dataset.")
            else:
                st.info(f"ℹ️ {stats['atualizadas']} amostras atualizadas (nenhuma nova).")
            time.sleep(1.5)
            st.rerun()

if st.session_state['dataset_oleo'] is None:
    ui_empty_state("A aguardar upload dos laudos para rodar o algoritmo e gerar relatórios.", icon="🔬")
    st.stop()

df = st.session_state['dataset_oleo'].copy()

# ==============================================================================
# BARRA LATERAL: FILTROS DE BUSCA
# ==============================================================================
with st.sidebar:
    st.header("🔍 Filtros de Análise")

    if 'DATA_COLETA' in df.columns:
        datas_validas = df['DATA_COLETA'].dropna()
        if not datas_validas.empty:
            min_d = datas_validas.min().date()
            max_d = datas_validas.max().date()
            datas = st.date_input("Período da Coleta", [min_d, max_d])
        else:
            datas = None
    else:
        datas = None

    filtro_status = st.radio("Condição da Amostra:", ["🚨 Apenas Críticas/Alertas", "Todas", "✅ Apenas Normais"],
                             index=1)

    if 'Família do equipamento' in df.columns:
        familias = sorted(df['Família do equipamento'].dropna().astype(str).unique().tolist())
        filtro_familias = st.multiselect("Família de Máquina:", options=familias,
                                          help="Filtra por categoria de equipamento")
    else:
        filtro_familias = []

    # Filtro por Frota específica — útil para gerar PDF de uma máquina específica
    if 'FROTA' in df.columns:
        # Se houver família selecionada, mostra apenas frotas dessas famílias
        if filtro_familias:
            df_frotas_disp = df[df['Família do equipamento'].astype(str).isin(filtro_familias)]
        else:
            df_frotas_disp = df
        frotas_disponiveis = sorted(
            df_frotas_disp['FROTA'].dropna().astype(str).unique().tolist()
        )
        filtro_frotas = st.multiselect(
            f"Frota específica ({len(frotas_disponiveis)} disponíveis):",
            options=frotas_disponiveis,
            help="Selecione uma ou mais frotas para gerar relatório apenas dessas máquinas. "
                 "Deixe vazio para incluir todas."
        )
    else:
        filtro_frotas = []

# --- APLICAÇÃO DOS FILTROS ---
df_view = df.copy()

if datas and len(datas) == 2:
    d_inicio = pd.to_datetime(datas[0])
    d_fim = pd.to_datetime(datas[1])
    df_view = df_view[(df_view['DATA_COLETA'] >= d_inicio) & (df_view['DATA_COLETA'] <= d_fim)]

if filtro_status == "🚨 Apenas Críticas/Alertas":
    df_view = df_view[df_view['STATUS_CORRIGIDO'].isin(['Crítico', 'Alerta'])]
elif filtro_status == "✅ Apenas Normais":
    df_view = df_view[df_view['STATUS_CORRIGIDO'] == 'Normal']

if filtro_familias:
    df_view = df_view[df_view['Família do equipamento'].isin(filtro_familias)]

if filtro_frotas:
    df_view = df_view[df_view['FROTA'].astype(str).isin(filtro_frotas)]

# Banner de filtros ativos — importante para o usuário saber o escopo do PDF
filtros_ativos = []
if filtro_familias:
    filtros_ativos.append(f"**Família:** {', '.join(filtro_familias)}")
if filtro_frotas:
    if len(filtro_frotas) <= 5:
        filtros_ativos.append(f"**Frota(s):** {', '.join(filtro_frotas)}")
    else:
        filtros_ativos.append(f"**{len(filtro_frotas)} frotas selecionadas**")
if filtro_status != "Todas":
    filtros_ativos.append(f"**Status:** {filtro_status}")
if datas and len(datas) == 2:
    filtros_ativos.append(f"**Período:** {datas[0].strftime('%d/%m/%Y')} a {datas[1].strftime('%d/%m/%Y')}")

if filtros_ativos:
    st.warning(
        f"🎯 **Filtros aplicados — o PDF e o Excel serão gerados apenas com este subconjunto:** "
        + "  |  ".join(filtros_ativos),
        icon="🎯"
    )

st.markdown("---")

# --- KPIs GERAIS ---
total_amostras = len(df_view)
criticas = len(df_view[df_view['STATUS_CORRIGIDO'] == 'Crítico'])
alertas = len(df_view[df_view['STATUS_CORRIGIDO'] == 'Alerta'])
normais = len(df_view[df_view['STATUS_CORRIGIDO'] == 'Normal'])

pct_normais = 0.0;
pct_alertas = 0.0;
pct_criticas = 0.0

if total_amostras > 0:
    pct_normais = round((normais / total_amostras) * 100, 1)
    pct_alertas = round((alertas / total_amostras) * 100, 1)
    pct_criticas = round((criticas / total_amostras) * 100, 1)

    diff = round(100.0 - (pct_normais + pct_alertas + pct_criticas), 1)
    if diff != 0:
        if normais >= alertas and normais >= criticas:
            pct_normais = round(pct_normais + diff, 1)
        elif alertas >= criticas:
            pct_alertas = round(pct_alertas + diff, 1)
        else:
            pct_criticas = round(pct_criticas + diff, 1)

c1, c2, c3, c4, c5 = st.columns(5)
ui_kpi_card(c1, "Amostras no Filtro", f"{total_amostras}", "🧪", "#3B82F6", "Volume analisado")

# Health Score médio com cor por faixa
health_medio = df_view['HEALTH_SCORE'].mean() if 'HEALTH_SCORE' in df_view.columns and len(df_view) > 0 else 0
if health_medio >= 75:
    cor_health = "#10B981"; rotulo_health = "Frota saudável"
elif health_medio >= 55:
    cor_health = "#F59E0B"; rotulo_health = "Atenção geral"
else:
    cor_health = "#EF4444"; rotulo_health = "Frota crítica"
ui_kpi_card(c2, "Health Score Médio", f"{health_medio:.0f}/100", "💚", cor_health, rotulo_health)

ui_kpi_card(c3, "Em Alerta", f"{alertas} ({pct_alertas:.1f}%)", "⚠️", "#F59E0B", "Atenção nas próximas trocas")
ui_kpi_card(c4, "Estado Crítico", f"{criticas} ({pct_criticas:.1f}%)", "🚨", "#EF4444" if criticas > 0 else "#10B981",
            "Exigem plano de ação")

# Top RPN — amostra de maior risco priorizado
if 'RPN' in df_view.columns and len(df_view) > 0:
    top_rpn = df_view.nlargest(1, 'RPN').iloc[0]
    rpn_label = f"#{top_rpn.get('FROTA', '?')} · RPN {top_rpn.get('RPN', 0):.0f}"
    rpn_desc = f"{top_rpn.get('COMPARTIMENTO', '-')[:18]}"
    ui_kpi_card(c5, "Maior Risco Priorizado", rpn_label, "🎯", "#DC2626", rpn_desc)
else:
    ui_kpi_card(c5, "Maior Risco Priorizado", "—", "🎯", "#94A3B8", "Sem dados")

st.markdown("<br>", unsafe_allow_html=True)

# ==============================================================================
# ABAS DE VISUALIZAÇÃO E FECHO DE CICLO
# ==============================================================================
tab_executivo, tab_tabela, tab_ia, tab_quimica, tab_saude, tab_tendencia, tab_feedback = st.tabs(
    ["🎯 Visão Executiva",
     "📋 Resumo de Pareceres",
     "🤖 Diagnóstico Sugerido",
     "🧪 Química Detalhada",
     "💧 Saúde do Óleo (FTIR / TBN)",
     "📈 Evolução Histórica (Timeline)",
     "🔄 Fecho de Ciclo (Gestão)"]
)

# ==============================================================================
# VISÃO EXECUTIVA — Espelha a capa do PDF
# ==============================================================================
with tab_executivo:
    st.markdown("##### 🎯 Sumário para a Diretoria")
    st.caption("As 10 amostras com maior risco priorizado (RPN) — esta é a lista para a reunião de manutenção.")

    if not df_view.empty and 'RPN' in df_view.columns:
        top10 = df_view.nlargest(10, 'RPN').copy()
        top10['#'] = range(1, len(top10) + 1)

        # Tabela com cor por faixa
        col_top = ['#', 'FROTA', 'COMPARTIMENTO', 'Família do equipamento',
                   'HEALTH_SCORE', 'RPN', 'TENDENCIA', 'ELEMENTOS_EM_ALERTA_ALS', 'DIAGNOSTICO_IA']
        col_top = [c for c in col_top if c in top10.columns]

        def cor_rpn(val):
            try:
                v = float(val)
                if v >= 80:
                    return "background-color: #FECACA; color: #7F1D1D; font-weight: bold;"
                elif v >= 40:
                    return "background-color: #FEF08A; color: #854D0E; font-weight: bold;"
            except:
                pass
            return ""

        def cor_health(val):
            try:
                v = float(val)
                if v < 40:
                    return "background-color: #FECACA; color: #7F1D1D; font-weight: bold;"
                elif v < 60:
                    return "background-color: #FEF08A; color: #854D0E;"
                elif v >= 75:
                    return "background-color: #DCFCE7; color: #166534;"
            except:
                pass
            return ""

        styled = top10[col_top].style
        if 'RPN' in col_top:
            styled = (styled.applymap if hasattr(styled, 'applymap') else styled.map)(cor_rpn, subset=['RPN'])
        if 'HEALTH_SCORE' in col_top:
            styled = (styled.applymap if hasattr(styled, 'applymap') else styled.map)(cor_health, subset=['HEALTH_SCORE'])

        st.dataframe(styled,
                      column_config={
                          "#": st.column_config.NumberColumn("#", width="small"),
                          "FROTA": "Frota",
                          "COMPARTIMENTO": "Compartimento",
                          "Família do equipamento": "Família",
                          "HEALTH_SCORE": st.column_config.NumberColumn("Health", format="%.0f", width="small"),
                          "RPN": st.column_config.NumberColumn("RPN", format="%.0f", width="small"),
                          "TENDENCIA": "Tendência",
                          "ELEMENTOS_EM_ALERTA_ALS": "🚩 ALS sinalizou",
                          "DIAGNOSTICO_IA": st.column_config.TextColumn("Causa raiz provável", width="large"),
                      },
                      hide_index=True, use_container_width=True)

        st.markdown("---")
        st.markdown("###### 📊 Distribuição da Saúde da Frota")
        st.caption(
            "Cada amostra recebe uma nota de 0 a 100 baseada na química do óleo "
            "e na classificação do laboratório ALS. **Quanto maior, melhor.**"
        )

        # Contagens por faixa
        n_exc  = int((df_view['HEALTH_SCORE'] >= 80).sum())
        n_bom  = int(((df_view['HEALTH_SCORE'] >= 60) & (df_view['HEALTH_SCORE'] < 80)).sum())
        n_at   = int(((df_view['HEALTH_SCORE'] >= 40) & (df_view['HEALTH_SCORE'] < 60)).sum())
        n_crit = int((df_view['HEALTH_SCORE'] < 40).sum())
        n_tot  = max(1, len(df_view))

        # Barra empilhada horizontal (visual rápido)
        fig_dist = go.Figure()
        for label, n, cor in [
            ('A · Excelente', n_exc,  '#16A34A'),
            ('B · Bom',       n_bom,  '#86EFAC'),
            ('C · Atenção',   n_at,   '#F59E0B'),
            ('D · Crítico',   n_crit, '#DC2626'),
        ]:
            if n > 0:
                fig_dist.add_trace(go.Bar(
                    x=[n], y=[''], orientation='h',
                    marker=dict(color=cor),
                    name=label,
                    text=f"{n} ({100*n/n_tot:.0f}%)",
                    textposition='inside',
                    insidetextfont=dict(color='white', size=14),
                    hovertemplate=f"<b>{label}</b><br>%{{x}} amostras<extra></extra>",
                ))
        fig_dist.update_layout(
            barmode='stack', height=110,
            margin=dict(t=10, b=10, l=10, r=10),
            xaxis=dict(showticklabels=False, showgrid=False),
            yaxis=dict(showticklabels=False),
            legend=dict(orientation='h', y=-0.4, x=0.5, xanchor='center'),
            showlegend=True,
        )
        st.plotly_chart(fig_dist, use_container_width=True)

        # Cards autoexplicativos — 4 colunas
        st.markdown(" ")  # espaço
        col_a, col_b, col_c, col_d = st.columns(4)

        def _card_faixa(col, letra, label, faixa, n, pct, desc, cor_fundo, cor_borda, cor_texto):
            col.markdown(
                f"""
                <div style="
                    background:{cor_fundo};
                    border-left:5px solid {cor_borda};
                    border-radius:6px;
                    padding:12px 14px;
                    min-height:130px;
                ">
                    <div style="display:flex; align-items:center; gap:8px; margin-bottom:6px;">
                        <span style="
                            background:{cor_borda}; color:#fff;
                            width:22px; height:22px; border-radius:50%;
                            display:flex; align-items:center; justify-content:center;
                            font-weight:bold; font-size:13px;
                        ">{letra}</span>
                        <strong style="color:{cor_texto}; font-size:13px;">{label}</strong>
                    </div>
                    <div style="color:{cor_texto}; font-size:11px; margin-bottom:4px;">
                        Nota {faixa}
                    </div>
                    <div style="color:{cor_borda}; font-size:18px; font-weight:bold; margin-bottom:6px;">
                        {n} <span style="font-size:12px; font-weight:normal;">({pct:.0f}%)</span>
                    </div>
                    <div style="color:{cor_texto}; font-size:11px; line-height:1.3;">
                        {desc}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        _card_faixa(col_a, 'A', 'EXCELENTE', '80 a 100', n_exc,  100*n_exc/n_tot,
                     'Equipamento operando dentro dos parâmetros normais. Nenhuma ação necessária no momento.',
                     '#DCFCE7', '#16A34A', '#166534')
        _card_faixa(col_b, 'B', 'BOM', '60 a 79', n_bom, 100*n_bom/n_tot,
                     'Pequenos sinais de desgaste dentro do esperado. Manter monitoramento nas próximas coletas.',
                     '#FEF3C7', '#86EFAC', '#15803D')
        _card_faixa(col_c, 'C', 'ATENÇÃO', '40 a 59', n_at, 100*n_at/n_tot,
                     'Anomalia confirmada pelo laboratório. Programar inspeção na próxima parada agendada.',
                     '#FED7AA', '#F59E0B', '#7C2D12')
        _card_faixa(col_d, 'D', 'CRÍTICO', '0 a 39', n_crit, 100*n_crit/n_tot,
                     'Risco operacional alto. Intervenção imediata recomendada — não opere a máquina sem inspeção.',
                     '#FECACA', '#DC2626', '#7F1D1D')

        # Tendências em bloco separado abaixo dos cards
        if 'TENDENCIA' in df_view.columns:
            st.markdown(" ")
            st.markdown("**Tendências detectadas nas amostras** _(comparando com a coleta anterior do mesmo compartimento)_")
            tend = df_view['TENDENCIA'].value_counts()
            tend_cols = st.columns(min(5, len(tend) or 1))
            for i, (t, n) in enumerate(tend.items()):
                emoji = ('🟢' if 'estabilizado' in t else
                         '🟡' if 'normal' in t else
                         '🟠' if 'acelerado' in t else
                         '🔴' if 'explosivo' in t else
                         '⚪')
                tend_cols[i % len(tend_cols)].markdown(f"{emoji} **{t}**\n\n{n} amostras")

        # Benchmark por fabricante
        if 'FABRICANTE' in df_view.columns:
            st.markdown("---")
            st.markdown("###### 🏭 Benchmark de Saúde por Fabricante")
            bench = (df_view.groupby('FABRICANTE')
                       .agg(Health_Médio=('HEALTH_SCORE', 'mean'),
                            Amostras=('HEALTH_SCORE', 'count'),
                            Críticos=('STATUS_CORRIGIDO', lambda s: (s == 'Crítico').sum()))
                       .reset_index()
                       .sort_values('Health_Médio'))
            bench = bench[bench['FABRICANTE'].astype(str).str.strip() != '-']

            if not bench.empty:
                fig_bench = px.bar(
                    bench, x='Health_Médio', y='FABRICANTE',
                    orientation='h',
                    color='Health_Médio',
                    color_continuous_scale=[(0, '#DC2626'), (0.4, '#F59E0B'), (0.7, '#86EFAC'), (1, '#16A34A')],
                    range_color=[0, 100],
                    text='Amostras',
                    title='Saúde média da frota por fabricante (com n de amostras)',
                )
                fig_bench.update_traces(texttemplate='n=%{text}', textposition='outside')
                fig_bench.update_layout(height=max(280, 40 * len(bench)),
                                         margin=dict(t=40, b=10, l=10, r=10),
                                         coloraxis_showscale=False)
                st.plotly_chart(fig_bench, use_container_width=True)
    else:
        st.info("Sem dados suficientes para o sumário executivo. Ajuste os filtros.")

with tab_tabela:
    st.markdown("##### Dashboards Analíticos")
    st.caption("Visão sumarizada das amostras. Os detalhes completos de Ação estão disponíveis no PDF exportado.")

    if not df_view.empty:
        c_graf1, c_graf2 = st.columns(2)
        with c_graf1:
            fig_fam = px.histogram(df_view, y='Família do equipamento', color='STATUS_CORRIGIDO',
                                   orientation='h',
                                   color_discrete_map={'Normal': '#10B981', 'Alerta': '#F59E0B', 'Crítico': '#EF4444'},
                                   title="Amostras por Família de Equipamento")
            fig_fam.update_layout(yaxis_title="", margin=dict(t=40, b=10, l=10, r=10))
            st.plotly_chart(fig_fam, use_container_width=True)

        with c_graf2:
            if df_view['DATA_COLETA'].notna().sum() > 0:
                fig_time = px.histogram(df_view, x='DATA_COLETA', color='STATUS_CORRIGIDO',
                                        color_discrete_map={'Normal': '#10B981', 'Alerta': '#F59E0B',
                                                            'Crítico': '#EF4444'},
                                        title="Volume de Coletas no Tempo")
                fig_time.update_layout(xaxis_title="Data", yaxis_title="Qtd", margin=dict(t=40, b=10, l=10, r=10))
                st.plotly_chart(fig_time, use_container_width=True)

        st.markdown("###### 📑 Tabela Resumo")


        def format_status(val):
            if val == 'Crítico': return 'background-color: #FECACA; color: #991B1B; font-weight: bold;'
            if val == 'Alerta': return 'background-color: #FEF08A; color: #A16207; font-weight: bold;'
            return 'color: #166534;'


        col_view = ['FROTA', 'COMPARTIMENTO', 'DATA_COLETA', 'STATUS_CORRIGIDO', 'AVALIACAO']
        col_view = [c for c in col_view if c in df_view.columns]

        st.dataframe(
            df_view[col_view].style.applymap(format_status, subset=['STATUS_CORRIGIDO']) if hasattr(df_view.style,
                                                                                                    'applymap') else
            df_view[col_view].style.map(format_status, subset=['STATUS_CORRIGIDO']),
            column_config={
                "FROTA": st.column_config.TextColumn("Máquina", width="small"),
                "COMPARTIMENTO": "Compartimento",
                "DATA_COLETA": st.column_config.DateColumn("Data", format="DD/MM/YYYY"),
                "STATUS_CORRIGIDO": "Status",
                "AVALIACAO": st.column_config.TextColumn("Parecer Resumido", width="large")
            },
            hide_index=True, use_container_width=True, height=300
        )
    else:
        st.success("Não há resultados para os filtros selecionados.")

with tab_ia:
    st.markdown("##### 🌳 Diagnóstico Sugerido (Análise Cruzada)")
    st.caption("Cruzamento entre os elementos químicos, flags da ALS e nossa lógica de análise. As colunas em vermelho indicam elementos individualmente sinalizados pelo laboratório.")

    df_problemas = df_view[df_view['STATUS_CORRIGIDO'].isin(['Crítico', 'Alerta'])].copy()
    if not df_problemas.empty:
        def format_diagnostico(val):
            return f"background-color: #FEF2F2; color: #991B1B; font-weight: 500;" if '✅' not in str(val) else ""

        def format_als_alert(val):
            v = str(val).strip()
            if v and v.lower() != 'nan':
                return "background-color: #FEE2E2; color: #7F1D1D; font-weight: bold;"
            return ""

        col_view = ['FROTA', 'COMPARTIMENTO', 'STATUS_CORRIGIDO',
                    'ELEMENTOS_EM_ALERTA_ALS', 'DIAGNOSTICO_IA',
                    'DADOS_RELEVANTES', 'RESPONSAVEL_ALS']
        col_view = [c for c in col_view if c in df_problemas.columns]

        styled = df_problemas[col_view].style
        if 'DIAGNOSTICO_IA' in col_view:
            styled = (styled.applymap if hasattr(styled, 'applymap') else styled.map)(
                format_diagnostico, subset=['DIAGNOSTICO_IA']
            )
        if 'ELEMENTOS_EM_ALERTA_ALS' in col_view:
            styled = (styled.applymap if hasattr(styled, 'applymap') else styled.map)(
                format_als_alert, subset=['ELEMENTOS_EM_ALERTA_ALS']
            )

        st.dataframe(
            styled,
            column_config={
                "FROTA": st.column_config.TextColumn("Máquina", width="small"),
                "COMPARTIMENTO": "Compartimento",
                "STATUS_CORRIGIDO": "Veredito",
                "ELEMENTOS_EM_ALERTA_ALS": st.column_config.TextColumn("🚩 ALS sinalizou", width="small"),
                "DIAGNOSTICO_IA": st.column_config.TextColumn("Diagnóstico Sugerido (Causa Provável)", width="large"),
                "DADOS_RELEVANTES": st.column_config.TextColumn("Química Relevante", width="medium"),
                "RESPONSAVEL_ALS": st.column_config.TextColumn("Analista", width="small"),
            },
            hide_index=True, use_container_width=True
        )

        # KPIs derivados
        st.markdown("###### 📊 KPIs de Anomalias")
        col_kpi1, col_kpi2, col_kpi3 = st.columns(3)
        elementos_top = (
            df_problemas['ELEMENTOS_EM_ALERTA_ALS']
            .dropna().astype(str)
            .str.split(', ').explode().str.strip()
        )
        elementos_top = elementos_top[elementos_top != ''].value_counts()
        with col_kpi1:
            if not elementos_top.empty:
                st.metric("Elemento mais frequente", elementos_top.index[0],
                          f"{elementos_top.iloc[0]} amostras")
            else:
                st.metric("Elemento mais frequente", "—")
        with col_kpi2:
            n_distinct = df_problemas['FROTA'].nunique()
            st.metric("Frotas com anomalia", f"{n_distinct}")
        with col_kpi3:
            if 'FABRICANTE' in df_problemas.columns:
                fabr_top = df_problemas['FABRICANTE'].value_counts()
                if not fabr_top.empty:
                    st.metric("Fabricante mais afetado", fabr_top.index[0],
                              f"{fabr_top.iloc[0]} amostras")
    else:
        st.info("Nenhuma anomalia crítica detetada neste filtro.")

with tab_quimica:
    st.markdown("##### 🧪 Química Detalhada por Categoria")
    st.caption("Visualização agrupada de todos os 22+ elementos analisados pela ALS. Selecione uma categoria abaixo.")

    if not df_view.empty:
        categoria = st.radio(
            "Categoria de elementos:",
            ["🔧 Metais de Desgaste", "🌪️ Contaminantes", "⚗️ Aditivos do Óleo", "📐 Condição Físico-Química"],
            horizontal=True
        )

        grupos_cat = {
            "🔧 Metais de Desgaste": ['Ferro', 'Cromo', 'Níquel', 'Chumbo', 'Estanho',
                                       'Cobre', 'Alumínio', 'Manganês', 'Titânio', 'Vanádio'],
            "🌪️ Contaminantes":     ['Silício', 'Sódio', 'Potássio', 'Boro'],
            "⚗️ Aditivos do Óleo":   ['Magnésio', 'Cálcio', 'Fósforo', 'Zinco', 'Bário', 'Molibdênio'],
            "📐 Condição Físico-Química": ['Viscosidade', 'TBN', 'KF_Agua', 'INDICE_PQ', 'Diluição Diesel'],
        }
        cols_cat = [c for c in grupos_cat[categoria] if c in df_view.columns]

        if cols_cat:
            # Tabela com cores por status
            cols_display = ['FROTA', 'COMPARTIMENTO', 'DATA_COLETA', 'STATUS_CORRIGIDO'] + cols_cat
            cols_display = [c for c in cols_display if c in df_view.columns]

            df_q = df_view[cols_display].copy()
            df_q = df_q.sort_values('STATUS_CORRIGIDO', key=lambda s: s.map(
                {'Crítico': 0, 'Alerta': 1, 'Normal': 2}
            ))

            def cor_celula_status(val, col_name):
                if col_name not in LIMITES_REFERENCIA:
                    return ""
                try:
                    v = float(val)
                except (ValueError, TypeError):
                    return ""
                limite_v, limite_a = LIMITES_REFERENCIA[col_name]
                if col_name == 'TBN':
                    if v >= limite_v: return "background-color: #DCFCE7"
                    if v >= limite_a: return "background-color: #FEF08A"
                    return "background-color: #FECACA; font-weight: bold"
                if v <= limite_v: return "background-color: #DCFCE7"
                if v <= limite_a: return "background-color: #FEF08A"
                return "background-color: #FECACA; font-weight: bold"

            styled_q = df_q.style
            for col in cols_cat:
                if col in LIMITES_REFERENCIA:
                    styled_q = (styled_q.applymap if hasattr(styled_q, 'applymap') else styled_q.map)(
                        lambda v, c=col: cor_celula_status(v, c), subset=[col]
                    )

            st.dataframe(styled_q,
                          column_config={
                              "DATA_COLETA": st.column_config.DateColumn("Data", format="DD/MM/YYYY"),
                              "STATUS_CORRIGIDO": "Status",
                          },
                          hide_index=True, use_container_width=True, height=420)

            # Gráfico de barras por elemento
            st.markdown("###### 📊 Comparativo entre amostras")
            elem_sel = st.selectbox("Elemento para visualizar:", cols_cat)
            if elem_sel and elem_sel in df_view.columns:
                df_plot = df_view[df_view[elem_sel] > 0].copy()
                if not df_plot.empty:
                    df_plot['rotulo'] = df_plot['FROTA'].astype(str) + ' - ' + df_plot['COMPARTIMENTO'].astype(str)
                    fig_el = px.bar(
                        df_plot.sort_values(elem_sel, ascending=False).head(20),
                        x='rotulo', y=elem_sel, color='STATUS_CORRIGIDO',
                        color_discrete_map={'Normal': '#10B981', 'Alerta': '#F59E0B', 'Crítico': '#EF4444'},
                        title=f"Top 20 amostras — {elem_sel}",
                        labels={'rotulo': 'Frota - Compartimento', elem_sel: f'{elem_sel} (ppm)'}
                    )
                    # Linhas de referência
                    if elem_sel in LIMITES_REFERENCIA:
                        lv, la = LIMITES_REFERENCIA[elem_sel]
                        fig_el.add_hline(y=lv, line_dash="dash", line_color="#F59E0B",
                                          annotation_text=f"Atenção ({lv})", annotation_position="right")
                        fig_el.add_hline(y=la, line_dash="dash", line_color="#EF4444",
                                          annotation_text=f"Crítico ({la})", annotation_position="right")
                    fig_el.update_layout(xaxis_tickangle=-45, height=450, margin=dict(t=40, b=120))
                    st.plotly_chart(fig_el, use_container_width=True)
                else:
                    st.info(f"Nenhuma amostra com valor > 0 para {elem_sel} no filtro atual.")
    else:
        st.info("Ajuste os filtros para visualizar os dados.")

with tab_saude:
    st.markdown("##### 💧 Saúde Físico-Química do Óleo")
    st.caption("Indicadores de degradação do óleo (TBN, viscosidade, FTIR, água). Permitem decidir quando trocar.")

    if not df_view.empty:
        # KPIs de saúde do óleo
        col_a, col_b, col_c, col_d = st.columns(4)

        if 'TBN' in df_view.columns:
            tbn_critico = df_view[(df_view['TBN'] > 0) & (df_view['TBN'] < 5)].shape[0]
            col_a.metric("TBN crítico (< 5)", f"{tbn_critico}", help="Óleo com poder de neutralização esgotado.")
        if 'KF_Agua' in df_view.columns:
            agua_alta = df_view[df_view['KF_Agua'] > 500].shape[0]
            col_b.metric("Água > 500 ppm", f"{agua_alta}", help="Risco de corrosão.")
        if 'Oxidação_FTIR' in df_view.columns:
            ox_alto = df_view[df_view['Oxidação_FTIR'] > 25].shape[0]
            col_c.metric("Oxidação crítica", f"{ox_alto}", help="Óleo degradado termicamente.")
        if 'Diluição Diesel' in df_view.columns:
            dd_alto = df_view[df_view['Diluição Diesel'] > 4].shape[0]
            col_d.metric("Diluição > 4%", f"{dd_alto}", help="Combustível no cárter.")

        st.markdown("---")
        st.markdown("###### Distribuição dos indicadores")

        indicadores = [
            ('TBN', 'TBN (mgKOH/g) — quanto MAIOR, melhor'),
            ('Viscosidade', 'Viscosidade @100°C (cSt)'),
            ('KF_Agua', 'Água por Karl Fischer (ppm)'),
            ('Oxidação_FTIR', 'Oxidação (FTIR)'),
            ('Nitração_FTIR', 'Nitração (FTIR)'),
            ('Sulfatação_FTIR', 'Sulfatação (FTIR)'),
        ]
        cols_disp = [(k, t) for k, t in indicadores if k in df_view.columns and df_view[k].sum() > 0]

        if cols_disp:
            ind_sel = st.selectbox("Indicador para análise:",
                                    options=[k for k, _ in cols_disp],
                                    format_func=lambda x: dict(cols_disp).get(x, x))

            df_ind = df_view[df_view[ind_sel] > 0].copy()
            if not df_ind.empty:
                col_g1, col_g2 = st.columns([2, 1])
                with col_g1:
                    fig_hist = px.histogram(df_ind, x=ind_sel, color='STATUS_CORRIGIDO',
                                             nbins=20,
                                             color_discrete_map={'Normal': '#10B981',
                                                                  'Alerta': '#F59E0B',
                                                                  'Crítico': '#EF4444'},
                                             title=f"Distribuição — {ind_sel}")
                    if ind_sel in LIMITES_REFERENCIA:
                        lv, la = LIMITES_REFERENCIA[ind_sel]
                        fig_hist.add_vline(x=lv, line_dash="dash", line_color="#F59E0B")
                        fig_hist.add_vline(x=la, line_dash="dash", line_color="#EF4444")
                    fig_hist.update_layout(height=380, margin=dict(t=40, b=10))
                    st.plotly_chart(fig_hist, use_container_width=True)

                with col_g2:
                    st.markdown("**Estatísticas**")
                    st.metric("Média", f"{df_ind[ind_sel].mean():.2f}")
                    st.metric("Mediana", f"{df_ind[ind_sel].median():.2f}")
                    st.metric("Máximo", f"{df_ind[ind_sel].max():.2f}")
                    st.metric("Amostras", f"{len(df_ind)}")
        else:
            st.info("Nenhum indicador de saúde do óleo disponível neste filtro.")
    else:
        st.info("Ajuste os filtros para visualizar os dados.")

with tab_tendencia:
    st.markdown("##### 📈 Evolução Histórica e Benchmarking")
    st.caption("Analise o desgaste de uma frota ou cruze os dados de várias máquinas para comparação (Benchmarking).")

    if not df_view.empty and 'FROTA' in df_view.columns:
        frotas_disponiveis = sorted(df_view['FROTA'].dropna().unique().tolist())
        # --- ATUALIZADO: Seleção Múltipla para Benchmarking ---
        frotas_selecionadas = st.multiselect("Selecione a(s) Frota(s) para análise e comparação:",
                                             options=frotas_disponiveis)

        if frotas_selecionadas:
            df_frota = df_view[df_view['FROTA'].isin(frotas_selecionadas)].copy()
            df_frota = df_frota.sort_values(by='DATA_COLETA')

            if not df_frota.empty:
                # --- ATUALIZADO: Alerta de Degradação Acelerada Matemática ---
                elementos_criticos = ['Ferro', 'Silício', 'Cobre', 'Chumbo', 'Alumínio']
                alertas_globais_degradacao = []

                for frota in frotas_selecionadas:
                    comps = df_frota[df_frota['FROTA'] == frota]['COMPARTIMENTO'].unique()
                    for comp in comps:
                        df_sub = df_frota[
                            (df_frota['FROTA'] == frota) & (df_frota['COMPARTIMENTO'] == comp)].sort_values(
                            'DATA_COLETA')
                        if len(df_sub) >= 2:
                            last = df_sub.iloc[-1]
                            prev = df_sub.iloc[-2]
                            for el in elementos_criticos:
                                if el in df_sub.columns:
                                    if last[el] > 15 and prev[el] > 0 and last[el] > (prev[el] * 1.5):
                                        alertas_globais_degradacao.append(
                                            f"**{frota} ({comp}):** {el} subiu de {prev[el]} para {last[el]} (+{((last[el] / prev[el]) - 1) * 100:.0f}%)")
                                    elif last[el] > 15 and prev[el] == 0:
                                        alertas_globais_degradacao.append(
                                            f"**{frota} ({comp}):** {el} disparou de 0 para {last[el]}")

                if alertas_globais_degradacao:
                    st.error("🚨 **ALERTA DE DEGRADAÇÃO ACELERADA DETECTADA** nas últimas amostras:\n" + "\n".join(
                        [f"- {a}" for a in alertas_globais_degradacao]))

                # --- Análise de Volatilidade ---
                mudancas_status = (df_frota['STATUS_CORRIGIDO'] != df_frota['STATUS_CORRIGIDO'].shift()).sum() - 1
                mudancas_status = max(0, mudancas_status)

                cor_volatilidade = "green" if mudancas_status <= 1 else ("orange" if mudancas_status <= 3 else "red")
                if len(frotas_selecionadas) == 1:
                    st.markdown(
                        f"**Índice de Estabilidade:** <span style='color:{cor_volatilidade}'>Mudou de estado {mudancas_status} vezes no histórico analisado.</span>",
                        unsafe_allow_html=True)

                # --- Gráfico de Timeline de Status (Saúde) ---
                df_frota['STATUS_VALOR'] = df_frota['STATUS_CORRIGIDO'].map({'Normal': 1, 'Alerta': 2, 'Crítico': 3})
                df_frota['FROTA_COMPARTIMENTO'] = df_frota['FROTA'] + " - " + df_frota['COMPARTIMENTO']
                df_frota['MARCADOR'] = df_frota['TROCOU_OLEO_FLAG'].apply(
                    lambda x: 'diamond' if x == 'Sim' else 'circle')

                fig_status = px.scatter(
                    df_frota, x='DATA_COLETA', y='STATUS_CORRIGIDO', color='STATUS_CORRIGIDO',
                    color_discrete_map={'Normal': '#10B981', 'Alerta': '#F59E0B', 'Crítico': '#EF4444'},
                    symbol='TROCOU_OLEO_FLAG', symbol_sequence=['circle', 'diamond'],
                    hover_data=['FROTA_COMPARTIMENTO', 'AVALIACAO'],
                    title=f"Linha do Tempo de Saúde",
                    labels={'DATA_COLETA': 'Data', 'STATUS_CORRIGIDO': 'Estado',
                            'TROCOU_OLEO_FLAG': 'Houve Troca de Óleo?'}
                )

                for f_c in df_frota['FROTA_COMPARTIMENTO'].unique():
                    df_linha = df_frota[df_frota['FROTA_COMPARTIMENTO'] == f_c]
                    fig_status.add_trace(go.Scatter(
                        x=df_linha['DATA_COLETA'], y=df_linha['STATUS_CORRIGIDO'],
                        mode='lines', line=dict(color='rgba(150, 150, 150, 0.4)', width=2),
                        name=f_c, hoverinfo='skip', showlegend=False
                    ))

                fig_status.update_traces(marker=dict(size=12, line=dict(width=1, color='DarkSlateGrey')))
                fig_status.update_yaxes(categoryorder='array', categoryarray=['Normal', 'Alerta', 'Crítico'])
                st.plotly_chart(fig_status, use_container_width=True)

                st.markdown("---")

                # --- Gráfico de Linha PPM (Química) ---
                elementos_quimicos = ['Ferro', 'Silício', 'Cobre', 'Alumínio', 'Cromo', 'Chumbo', 'Viscosidade',
                                      'Diluição Diesel', 'INDICE_PQ', 'Potássio', 'Molibdênio', 'Fuligem']
                elementos_disponiveis = [e for e in elementos_quimicos if
                                         e in df_frota.columns and df_frota[e].sum() > 0]

                if elementos_disponiveis:
                    elementos_selecionados = st.multiselect(
                        "Selecione os indicadores químicos para cruzar (em ppm / %):",
                        options=elementos_disponiveis,
                        default=[e for e in ['Ferro', 'Silício'] if
                                 e in elementos_disponiveis] or elementos_disponiveis[:1]
                    )

                    if elementos_selecionados:
                        fig_trend = px.line(
                            df_frota, x='DATA_COLETA', y=elementos_selecionados, color='FROTA_COMPARTIMENTO',
                            markers=True,
                            title=f"Evolução Química (PPM)",
                            labels={'value': 'Concentração', 'DATA_COLETA': 'Data da Coleta', 'variable': 'Indicador',
                                    'FROTA_COMPARTIMENTO': 'Equipamento'}
                        )
                        fig_trend.update_layout(hovermode="x unified")
                        st.plotly_chart(fig_trend, use_container_width=True)

                        st.markdown(f"###### Histórico Tabular Recente")
                        col_hist = ['DATA_COLETA', 'FROTA', 'COMPARTIMENTO', 'STATUS_CORRIGIDO',
                                    'TROCOU_OLEO_FLAG'] + elementos_selecionados
                        st.dataframe(df_frota[col_hist].sort_values(by='DATA_COLETA', ascending=False), hide_index=True,
                                     use_container_width=True)
                else:
                    st.warning("Nenhum dado numérico de elementos químicos encontrado para esta frota.")
            else:
                st.info("Sem dados temporais disponíveis.")
    else:
        st.info("Selecione uma ou mais frotas para analisar o histórico e a degradação.")

with tab_feedback:
    st.markdown("##### 🔄 Fecho de Ciclo e Geração de Ordem de Serviço")
    st.caption(
        "Marque a caixa 'Gerar O.S.' para abrir um ticket direto para a oficina e salvar no banco de dados da Manutenção!")

    df_feed_ui = df_view[df_view['STATUS_CORRIGIDO'].isin(['Crítico', 'Alerta'])].copy()

    if not df_feed_ui.empty:
        # Adiciona a caixa de seleção para a inteligência de OS
        df_feed_ui.insert(0, 'Gerar_OS', False)

        colunas_editaveis = ['Gerar_OS', 'NUM_AMOSTRA', 'FROTA', 'COMPARTIMENTO', 'STATUS_CORRIGIDO', 'ACAO_GESTAO',
                             'STATUS_ACAO']
        colunas_editaveis = [c for c in colunas_editaveis if c in df_feed_ui.columns]

        edited_df = st.data_editor(
            df_feed_ui[colunas_editaveis],
            column_config={
                "Gerar_OS": st.column_config.CheckboxColumn("⚙️ Criar O.S.?",
                                                            help="Abre uma Ordem de Serviço automática.",
                                                            default=False),
                "NUM_AMOSTRA": st.column_config.TextColumn("Amostra", disabled=True),
                "FROTA": st.column_config.TextColumn("Máquina", disabled=True),
                "COMPARTIMENTO": st.column_config.TextColumn("Local", disabled=True),
                "STATUS_CORRIGIDO": st.column_config.TextColumn("Status", disabled=True),
                "ACAO_GESTAO": st.column_config.TextColumn("📝 Observações / Ação", required=False, width="large"),
                "STATUS_ACAO": st.column_config.SelectboxColumn("Situação",
                                                                options=['Pendente', 'Em Andamento', 'Concluída'],
                                                                required=True)
            },
            hide_index=True, use_container_width=True, key="editor_feedback"
        )

        if st.button("💾 Executar Ações e Gerar O.S.", type="primary"):
            try:
                os_geradas = 0

                with db_connection() as conn:
                    cursor = conn.cursor()

                    # Verificação defensiva: tabelas podem não existir em ambientes novos
                    tabelas = {r[0] for r in cursor.execute(
                        "SELECT name FROM sqlite_master WHERE type='table'"
                    ).fetchall()}

                    map_eq = {}
                    map_op = {}

                    if 'equipamentos' in tabelas:
                        map_eq = {r[1].upper().strip(): r[0]
                                  for r in cursor.execute("SELECT id, frota FROM equipamentos")}

                    if 'tipos_operacao' in tabelas:
                        map_op = {r[1].upper().strip(): r[0]
                                  for r in cursor.execute("SELECT id, nome FROM tipos_operacao")}

                    # Sem fallback hardcoded para ID 1
                    op_id_preditiva = (
                        map_op.get('MECÂNICA')
                        or map_op.get('MECANICA')
                        or map_op.get('PREVENTIVA')
                        or (next(iter(map_op.values()), None))
                    )

                    if op_id_preditiva is None and 'ordens_servico' in tabelas:
                        st.warning("Nenhum tipo de operação cadastrado. Cadastre ao menos um para gerar O.S.")

                    for _, row in edited_df.iterrows():
                        amostra = str(row['NUM_AMOSTRA']).strip()
                        acao = str(row.get('ACAO_GESTAO', '')).strip()
                        status_acao = str(row.get('STATUS_ACAO', 'Pendente')).strip()

                        if row.get('Gerar_OS', False) and op_id_preditiva is not None:
                            f_key = str(row['FROTA']).upper().strip()
                            if f_key in map_eq:
                                eq_id = map_eq[f_key]
                                desc_os = (
                                    f"[GERADO VIA PREDITIVA DE ÓLEO]\n"
                                    f"Amostra: {amostra}\n"
                                    f"Compartimento: {row['COMPARTIMENTO']}\n"
                                    f"Diagnóstico Lab: {row['STATUS_CORRIGIDO']}\n"
                                    f"Ação da Gestão: {acao}"
                                )
                                prio_os = "Alta" if row['STATUS_CORRIGIDO'] == 'Crítico' else "Média"
                                cursor.execute("""
                                    INSERT INTO ordens_servico (data_hora, equipamento_id, descricao, tipo_operacao_id, status, prioridade, classificacao, maquina_parada)
                                    VALUES (?, ?, ?, ?, 'Pendente', ?, 'Preditiva', 0)
                                """, (datetime.now(), eq_id, desc_os, op_id_preditiva, prio_os))
                                nova_os_id = cursor.lastrowid
                                acao = f"✅ OS #{nova_os_id} Gerada Automática. " + acao
                                os_geradas += 1
                            else:
                                logger.warning("Frota '%s' não encontrada no cadastro de equipamentos.", f_key)

                        if amostra and amostra != '-':
                            cursor.execute(
                                "UPDATE analises_oleo_feedback SET acao_gestao = ?, status_acao = ? WHERE amostra = ?",
                                (acao, status_acao, amostra))

                if os_geradas > 0:
                    st.balloons()
                    st.success(f"Ações atualizadas com sucesso e {os_geradas} Ordens de Serviço enviadas para a oficina!")
                else:
                    st.toast("Ações de feedback atualizadas com sucesso!", icon="✅")

                st.session_state['dataset_oleo'] = sincronizar_amostras_bd(st.session_state['dataset_oleo'])
                time.sleep(1.5)
                st.rerun()

            except Exception as e:
                st.error(f"Erro ao salvar ações: {e}")
                logger.exception("Falha ao executar ações e gerar O.S.")
    else:
        st.success("Não há máquinas críticas a aguardar feedback no filtro atual.")

# ==============================================================================
# EXPORTAÇÃO (PLANOS DE AÇÃO)
# ==============================================================================
st.markdown("---")
st.markdown("### 📋 Geração de Planos de Ação (Exportação)")
st.caption("Gere os relatórios. Se preencheu a aba 'Fecho de Ciclo' acima, os PDFs já sairão preenchidos!")

tipo_exportacao = st.radio(
    "Quais amostras deseja incluir no relatório final?",
    ["🚨 Apenas Críticas e em Alerta (Focado em Ação Imediata)",
     "📑 Todas as Amostras do Filtro Atual (Histórico Completo)"],
    horizontal=True
)

if df_view.empty:
    st.info("Ajuste os filtros para gerar o relatório.")
else:
    if "Apenas" in tipo_exportacao:
        df_export = df_view[df_view['STATUS_CORRIGIDO'].isin(['Crítico', 'Alerta'])].copy()
    else:
        df_export = df_view.copy()

    if df_export.empty:
        st.success("Tudo limpo! Não há amostras no filtro atual para gerar o plano de ação.")
    else:
        c_pdf, c_excel = st.columns(2)

        with st.spinner("A compilar Gráficos Gerenciais e Ordens de Serviço..."):
            excel_bytes = gerar_excel_plano_acao(df_export)
            pdf_bytes = gerar_pdf_plano_acao(df_export, df)

        nome_arq = f"Plano_Acao_Oleo_{datetime.now().strftime('%d%m%Y')}"

        with c_pdf:
            st.download_button(label=f"📄 Caderno de Ações em PDF", data=pdf_bytes, file_name=f"{nome_arq}.pdf",
                               mime="application/pdf", type="primary", use_container_width=True)

        with c_excel:
            if excel_bytes:
                st.download_button(label=f"📊 Planilha de Ações em Excel", data=excel_bytes,
                                   file_name=f"{nome_arq}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   type="primary", use_container_width=True)
            else:
                st.error("Biblioteca 'openpyxl' não instalada para gerar o ficheiro Excel.")